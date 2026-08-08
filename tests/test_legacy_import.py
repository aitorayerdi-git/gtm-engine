from __future__ import annotations

import json
from datetime import date, timedelta
from hashlib import sha256
from pathlib import Path
from typing import cast

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from gtm_engine.cli import main
from gtm_engine.excel import load_excel_bundle, workbook_formula_cells
from gtm_engine.io import load_bundle
from gtm_engine.legacy_import import (
    LegacyImportError,
    import_legacy_workbook,
    read_legacy_workbook,
)


def _hash(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _sheet(workbook: Workbook, name: str) -> Worksheet:
    return cast(Worksheet, workbook.create_sheet(name))


def _synthetic_legacy(path: Path) -> Path:
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    setup = _sheet(workbook, "SETUP")
    setup.append([None] * 16)
    setup.append(["#", "Book Name", "Active?", None, None, "#", "Underlying"])
    setup.cell(3, 1, 1)
    setup.cell(3, 2, "BOOK1")
    setup.cell(3, 3, "YES")
    setup.cell(3, 6, 1)
    setup.cell(3, 7, "TTF DA")
    setup.cell(3, 8, "MWh")
    setup.cell(3, 9, "Metodología Heren")
    setup.cell(3, 10, "YES")
    setup.cell(4, 6, 2)
    setup.cell(4, 7, "TTF MA")
    setup.cell(4, 8, "MWh")
    setup.cell(4, 9, "Month Ahead")
    setup.cell(4, 10, "YES")

    initial = _sheet(workbook, "INITIAL POSITION")
    initial["B2"] = "Initial Market Date"
    initial["C2"] = date(2026, 6, 30)
    initial["A4"] = "BOOK"
    initial["B4"] = "Initial PnL"
    initial["A5"] = "BOOK1"
    initial["B5"] = 100

    initial_data = _sheet(workbook, "INITIAL POSITION DATA")
    initial_data.append(
        ["Initial Market Date", "BOOK", "Underlying", "Delivery Month", "Initial Exposure"]
    )
    initial_data.append([date(2026, 6, 30), "BOOK1", "TTF DA", date(2026, 7, 1), 100])
    initial_data.append([date(2026, 6, 30), "BOOK1", "TTF MA", date(2026, 7, 1), 0])

    process = _sheet(workbook, "PROCESS")
    process["C15"] = date(2026, 6, 30)
    process["D15"] = date(2026, 7, 2)

    calendar = _sheet(workbook, "CALENDAR")
    calendar.append(["Market calendar"])
    calendar.append([None])
    calendar.append(["Date", "Day", "Holiday", "Weekend", "Market Day?"])
    cursor = date(2026, 6, 1)
    final = date(2026, 8, 31)
    while cursor <= final:
        calendar.append(
            [cursor, cursor.strftime("%a"), None, None, "YES" if cursor.weekday() < 5 else None]
        )
        cursor += timedelta(days=1)

    trades = _sheet(workbook, "TRADES")
    trades.append(["Trades"])
    trades.append(
        [
            "Trade Date",
            "Book",
            "Underlying",
            "Comments",
            "BUY / SELL",
            "Start Date",
            "End Date",
            "Daily Qty",
            "Price",
        ]
    )
    trades.append(
        [
            date(2026, 7, 3),
            "BOOK1",
            "TTF DA",
            "after cutoff",
            "BUY",
            date(2026, 7, 6),
            date(2026, 7, 6),
            1,
            17,
        ]
    )
    trades.append(
        [
            date(2026, 7, 1),
            "BOOK1",
            "TTF DA",
            "audit note",
            "BUY",
            date(2026, 7, 2),
            date(2026, 7, 2),
            10,
            19,
        ]
    )

    simulation = _sheet(workbook, "SIMULATION TRADES")
    simulation["Q1"] = "OFF"
    simulation.cell(2, 1, "Trade Date")
    simulation.cell(2, 15, "Scenario")
    for column, value in enumerate(
        (
            date(2026, 7, 1),
            "BOOK1",
            "TTF DA",
            None,
            "SELL",
            date(2026, 7, 2),
            date(2026, 7, 2),
            5,
            18,
        ),
        start=1,
    ):
        simulation.cell(3, column, value)
    simulation.cell(3, 15, "WHAT-IF")

    ttf = _sheet(workbook, "TTF")
    ttf.cell(5, 5, date(2026, 7, 1))
    for row, (market_date, price) in enumerate(
        (
            (date(2026, 6, 30), 25),
            (date(2026, 7, 1), 26),
            (date(2026, 7, 2), 27),
        ),
        start=9,
    ):
        ttf.cell(row, 4, market_date)
        ttf.cell(row, 5, price)
        ttf.cell(row - 2, 45, market_date)
        ttf.cell(row - 2, 53, price + 10)

    for name in ("Brent Dated", "HH"):
        curve = _sheet(workbook, name)
        curve.cell(5, 5, date(2026, 7, 1))
    for name in ("PVB-TTF", "PEG-TTF"):
        spread = _sheet(workbook, name)
        spread.cell(2, 3, date(2026, 7, 1))
        if name == "PVB-TTF":
            for row, (market_date, spread_value) in enumerate(
                (
                    (date(2026, 6, 30), 1),
                    (date(2026, 7, 1), 2),
                    (date(2026, 7, 2), 3),
                ),
                start=3,
            ):
                spread.cell(row, 2, market_date)
                spread.cell(row, 3, spread_value)

    fixings = _sheet(workbook, "FIXING PRICES")
    fixings.cell(4, 2, "TTF DA")
    fixings.cell(4, 3, "TTF MA")
    fixings.cell(6, 1, date(2026, 7, 1))
    fixings.cell(6, 2, 20)
    fixings.cell(6, 3, 20)
    fixings.cell(7, 1, date(2026, 7, 2))
    fixings.cell(7, 2, 21)
    fixings.cell(7, 3, 21)

    costs = _sheet(workbook, "COSTS")
    costs.cell(3, 2, "BOOK1")
    costs.cell(5, 1, date(2026, 7, 1))
    costs.cell(5, 2, 5)

    foto = _sheet(workbook, "Foto FO")
    foto.cell(4, 18, "BOOK1")
    foto.cell(4, 20, "BOOK1")
    foto.cell(6, 2, date(2026, 7, 1))
    foto.cell(6, 18, 2)
    foto.cell(6, 20, 1)

    workbook.save(path)
    workbook.close()
    return path


def test_legacy_import_publishes_both_interfaces_and_audit(tmp_path: Path) -> None:
    source = _synthetic_legacy(tmp_path / "legacy.xlsm")
    before = _hash(source)
    imported = import_legacy_workbook(source, tmp_path / "imported")

    assert _hash(source) == before
    assert imported.report.status == "CREATED"
    assert imported.report.error_count == 0
    assert imported.report.extracted_row_counts == {
        "books": 1,
        "underlyings": 2,
        "calendar": 92,
        "initial_exposure": 1,
        "initial_pnl": 1,
        "curve_prices": 18,
        "fixing_prices": 4,
        "operating_flows": 1,
        "trades": 2,
    }
    assert imported.report.skipped_row_counts == {
        "actual_trades_after_cutoff": 1,
        "zero_initial_exposure": 1,
    }
    assert imported.bundle.trades[1].scenario == "WHAT-IF"
    prompt_prices = {
        (row.market_date, row.underlying, row.delivery_month): row.curve_price
        for row in imported.bundle.curve_prices
    }
    assert prompt_prices[(date(2026, 6, 30), "TTF DA", date(2026, 7, 1))] == 35
    assert prompt_prices[(date(2026, 7, 1), "TTF DA", date(2026, 7, 1))] == 36
    assert prompt_prices[(date(2026, 7, 2), "TTF DA", date(2026, 7, 1))] == 37
    assert prompt_prices[(date(2026, 6, 30), "Index PVB", date(2026, 7, 1))] == 36
    assert prompt_prices[(date(2026, 7, 1), "Index PVB", date(2026, 7, 1))] == 38
    assert prompt_prices[(date(2026, 7, 2), "Index PVB", date(2026, 7, 1))] == 40

    normalized = load_bundle(imported.normalized_bundle)
    excel = load_excel_bundle(imported.excel_workbook)
    assert normalized.model_copy(update={"input_hashes": {}}) == imported.bundle.model_copy(
        update={"input_hashes": {}}
    )
    assert excel.model_copy(update={"input_hashes": {}}) == imported.bundle.model_copy(
        update={"input_hashes": {}}
    )
    assert workbook_formula_cells(imported.excel_workbook) == ()

    audit = json.loads(imported.audit_json.read_text(encoding="utf-8"))
    assert audit["source"]["sha256"] == before
    assert audit["source"]["unchanged_during_import"] is True
    workbook = load_workbook(imported.excel_workbook, data_only=False, read_only=False)
    try:
        assert workbook["START HERE"]["B4"].value == "IMPORTED — NOT BUILT"
        assert "tblLegacyImportIssues" in workbook["LEGACY IMPORT"].tables
    finally:
        workbook.close()


def test_legacy_import_cli_and_fail_closed_errors(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = _synthetic_legacy(tmp_path / "legacy.xlsm")
    output = tmp_path / "cli-import"
    assert (
        main(
            [
                "legacy-import",
                "--workbook",
                str(source),
                "--output",
                str(output),
                "--historical-end",
                "2026-07-02",
            ]
        )
        == 0
    )
    message = json.loads(capsys.readouterr().out)
    assert message["status"] == "CREATED"
    assert Path(message["workbook"]).exists()

    assert main(["legacy-import", "--workbook", str(source), "--output", str(output)]) == 2
    failure = json.loads(capsys.readouterr().out)
    assert failure["status"] == "FAILED"
    assert "already exists" in failure["error"]

    broken = tmp_path / "broken.xlsx"
    workbook = Workbook()
    workbook.save(broken)
    workbook.close()
    with pytest.raises(LegacyImportError, match="missing sheet"):
        read_legacy_workbook(broken)


def test_real_legacy_workbook_inventory_when_available() -> None:
    source = Path("Gas_Trading_Model 070826.xlsm")
    if not source.exists():
        pytest.skip("Repository checkout does not include the private legacy workbook")
    before = _hash(source)
    bundle, report = read_legacy_workbook(source)

    assert _hash(source) == before
    assert report.source_sha256 == before
    assert len(bundle.books) == 13
    assert len(bundle.underlyings) == 18
    assert len(bundle.calendar) == 1097
    assert len(bundle.initial_exposure) == 101
    assert len(bundle.initial_pnl) == 13
    assert len(bundle.trades) == 476
    assert len(bundle.curve_prices) == 5430
    assert len(bundle.fixing_prices) == 572
    assert len(bundle.operating_flows) == 93
    assert {issue.code for issue in report.issues} >= {
        "FX_CONVERSION_NOT_DEFINED",
        "FIXING_PRICE_SERIES_EMPTY",
        "ZERO_DAILY_QTY",
        "ZERO_EXECUTION_PRICE",
    }
