from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from gtm_engine.excel import load_excel_bundle
from gtm_engine.pipeline import build
from scripts.create_fixing_test import main


def test_fixing_and_exposure_workbooks_rebuild_without_local_source(
    tmp_path: Path,
    monkeypatch,
) -> None:
    input_path = tmp_path / "Input synthetic fixing delivery Jul26.xlsx"
    output_path = tmp_path / "output synthetic fixing delivery Jul26.xlsx"
    mapping = Path(__file__).parents[1] / "docs" / "GTM_ACTIVE_SETUP_MAPPING_v0.3.csv"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "create_fixing_test.py",
            "--mapping",
            str(mapping),
            "--input",
            str(input_path),
            "--output",
            str(output_path),
            "--july-delivery",
        ],
    )

    assert main() == 0
    bundle = load_excel_bundle(input_path)
    result = build(bundle)
    assert result.manifest.status.value == "VERIFIED"
    assert not result.validation
    assert len(bundle.trades) == 17
    assert len(bundle.delivery_elections) == 2

    workbook = load_workbook(output_path, data_only=False, read_only=False)
    try:
        assert workbook.sheetnames == [
            "Fixing volume by fixing date",
            "Fixing PnL by fixing date",
            "Fixing PnL by market date",
            "Exposure data",
            "Lists",
            "Exposure by Market Date",
            "Delta Exposure MtM",
        ]
        for title in ("Exposure by Market Date", "Delta Exposure MtM"):
            sheet = workbook[title]
            assert sheet["A6"].value.year == 2026
            assert sheet["A6"].value.month == 1
            assert sheet["A41"].value.year == 2028
            assert sheet["A41"].value.month == 12
    finally:
        workbook.close()
