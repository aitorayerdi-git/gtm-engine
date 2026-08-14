from __future__ import annotations

import json
from datetime import date
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from gtm_engine.cli import main
from gtm_engine.excel import (
    SHEET_ORDER,
    ExcelAdapterError,
    build_excel_workbook,
    create_excel_template,
    load_excel_bundle,
    load_setup_mapping,
    workbook_formula_cells,
)
from gtm_engine.models import InitialExposure, Side, Trade, TradeSource

from .helpers import D, base_bundle, underlying, with_prices


def _hash(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _economic_bundle():
    trade = Trade(
        source_row_id="TRADE-EXCEL-1",
        trade_date=date(2026, 1, 5),
        book="BOOK1",
        underlying="GAS",
        side=Side.BUY,
        start_date=date(2026, 1, 7),
        end_date=date(2026, 1, 7),
        daily_qty=D("10"),
        execution_price=D("19.5"),
        trade_source=TradeSource.ACTUAL,
    )
    return with_prices(base_bundle(historical_end_date=date(2026, 1, 9), trades=(trade,)))


def _daily_report_bundle():
    trade = Trade(
        source_row_id="TRADE-REPORT-1",
        trade_date=date(2026, 1, 5),
        book="BOOK1",
        underlying="TTF DA",
        side=Side.BUY,
        start_date=date(2026, 1, 10),
        end_date=date(2026, 1, 10),
        daily_qty=D("10"),
        execution_price=D("19.5"),
        trade_source=TradeSource.ACTUAL,
    )
    return with_prices(
        base_bundle(
            historical_end_date=date(2026, 1, 9),
            underlyings=(underlying("TTF DA"),),
            trades=(trade,),
        )
    )


def _table_data_count(path: Path, sheet_name: str, table_name: str) -> int:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook[sheet_name]
        table = sheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        assert None not in (min_col, min_row, max_col, max_row)
        return sum(
            any(sheet.cell(row, column).value is not None for column in range(min_col, max_col + 1))
            for row in range(min_row + 1, max_row + 1)
        )
    finally:
        workbook.close()


def test_excel_template_round_trip_contract_and_setup_mapping(tmp_path: Path) -> None:
    mapping = Path("docs/GTM_ACTIVE_SETUP_MAPPING_v0.3.csv")
    books, underlyings = load_setup_mapping(mapping)
    assert len(books) == 13
    assert len(underlyings) == 17

    original = _economic_bundle()
    interface = create_excel_template(tmp_path / "interface.xlsx", bundle=original)
    before = _hash(interface)
    loaded = load_excel_bundle(interface)

    assert loaded.model_copy(update={"input_hashes": {}}) == original
    assert loaded.input_hashes == {"excel_workbook": before}
    assert _hash(interface) == before
    assert workbook_formula_cells(interface) == ()

    workbook = load_workbook(interface, data_only=False, read_only=False)
    try:
        assert tuple(workbook.sheetnames) == SHEET_ORDER
        assert workbook["START HERE"]["B4"].value == "NOT BUILT"
        assert workbook["CONTROL"]["B9"].value == "OFF"
        assert "tblManualDates" in workbook["MANUAL CHANGES"].tables
        assert (
            workbook["MANUAL CHANGES"]["A6"].value
            == "Market Date INITIAL POSITION PVB MO"
        )
        assert workbook["MANUAL CHANGES"]["B6"].value is None
        assert workbook["MANUAL CHANGES"]["B8"].value.date() == date(2026, 1, 9)
        assert workbook["TRADES"].freeze_panes == "A5"
        assert "tblTrades" in workbook["TRADES"].tables
        assert workbook["TRADES"].sheet_view.showGridLines is False
        assert all(sheet.auto_filter.ref is None for sheet in workbook.worksheets if sheet.tables)
    finally:
        workbook.close()


def test_excel_build_publishes_results_and_never_changes_source(tmp_path: Path) -> None:
    source = create_excel_template(tmp_path / "source.xlsx", bundle=_economic_bundle())
    before = _hash(source)
    output_root = tmp_path / "results"

    result, run_directory, result_workbook = build_excel_workbook(source, output_root)

    assert result.manifest.status.value == "PUBLISHED"
    assert result_workbook == run_directory / "GTM_Result.xlsx"
    assert result_workbook.exists()
    latest_path = output_root / "GTM_LATEST.xlsx"
    latest = load_workbook(latest_path, read_only=True, data_only=False)
    try:
        assert latest.sheetnames == [
            "Fixing volume by fixing date",
            "Fixing PnL by fixing date",
            "Fixing PnL by market date",
            "Exposure data",
            "Lists",
            "Exposure by Market Date",
            "Delta Exposure MtM",
        ]
    finally:
        latest.close()
    assert (run_directory / "GTM_Output.xlsx").read_bytes() == latest_path.read_bytes()
    assert _hash(source) == before
    assert workbook_formula_cells(result_workbook) == ()
    assert _table_data_count(result_workbook, "FIXINGS", "tblFixings") == len(result.fixings)
    assert _table_data_count(result_workbook, "EXPOSURE", "tblExposure") == len(result.exposure)
    assert _table_data_count(result_workbook, "DAILY PNL", "tblDailyPnl") == len(result.pnl)
    assert _table_data_count(result_workbook, "EVENT LEDGER", "tblEventLedger") == len(
        result.event_ledger
    )

    workbook = load_workbook(result_workbook, data_only=False, read_only=False)
    try:
        assert workbook["START HERE"]["B4"].value == "PUBLISHED"
        assert workbook["START HERE"]["B5"].value == result.manifest.build_id
        assert workbook["START HERE"]["B8"].value == 0
        assert workbook["BUILD MANIFEST"]["B8"].value is not None
        assert all(sheet.auto_filter.ref is None for sheet in workbook.worksheets if sheet.tables)
    finally:
        workbook.close()


def test_excel_build_adds_legacy_format_daily_report_for_last_two_market_dates(
    tmp_path: Path,
) -> None:
    source = create_excel_template(tmp_path / "source.xlsx", bundle=_daily_report_bundle())
    result, _, result_workbook = build_excel_workbook(source, tmp_path / "results")

    assert result.manifest.status.value == "PUBLISHED"
    workbook = load_workbook(result_workbook, data_only=False, read_only=False)
    try:
        assert "Daily Report D2" in workbook.sheetnames
        assert workbook.sheetnames.index("Daily Report D2") < workbook.sheetnames.index(
            "EVENT LEDGER"
        )
        report = workbook["Daily Report D2"]
        assert report["A1"].value == "Daily Report D2"
        assert report["C1"].value.date() == date(2026, 1, 8)
        assert report["E1"].value.date() == date(2026, 1, 9)

        assert report["A5"].value == "BOOK1"
        assert report["B5"].value == pytest.approx(-50)
        assert report["C5"].value == pytest.approx(-250)
        assert report["D5"].value == pytest.approx(200)

        assert report["A18"].value == "BOOK1"
        assert report["B18"].value.date() == date(2026, 1, 1)
        assert report["C18"].value == pytest.approx(-250)
        assert report["D18"].value == pytest.approx(200)

        assert report["B26"].value.date() == date(2026, 1, 1)
        assert report["E26"].value == pytest.approx(-250)
        assert report.auto_filter.ref is None
        assert report.sheet_view.showGridLines is True
    finally:
        workbook.close()

    assert workbook_formula_cells(result_workbook) == ()


def test_failed_excel_build_retains_diagnostics_and_preserves_latest(tmp_path: Path) -> None:
    output_root = tmp_path / "results"
    success_source = create_excel_template(tmp_path / "success.xlsx", bundle=_economic_bundle())
    success, _, _ = build_excel_workbook(success_source, output_root)
    assert success.manifest.status.value == "PUBLISHED"
    latest_before = _hash(output_root / "GTM_LATEST.xlsx")

    missing_prices = base_bundle(
        historical_end_date=date(2026, 1, 9),
        initial_exposure=(
            InitialExposure(
                initial_market_date=date(2026, 1, 2),
                book="BOOK1",
                underlying="GAS",
                delivery_month=date(2026, 1, 1),
                exposure_volume=D("100"),
                source_row_id="OPEN-MISSING-PRICE",
            ),
        ),
    )
    failed_source = create_excel_template(tmp_path / "failed-source.xlsx", bundle=missing_prices)
    failed, run_directory, failed_workbook = build_excel_workbook(failed_source, output_root)

    assert failed.manifest.status.value == "FAILED"
    assert failed_workbook == run_directory / "GTM_Failed.xlsx"
    assert run_directory.parent == output_root / "failed"
    assert _hash(output_root / "GTM_LATEST.xlsx") == latest_before
    assert _table_data_count(failed_workbook, "VALIDATION", "tblValidation") == len(
        failed.validation
    )
    assert _table_data_count(failed_workbook, "FIXINGS", "tblFixings") == 0
    workbook = load_workbook(failed_workbook, data_only=False, read_only=False)
    try:
        assert workbook["START HERE"]["B4"].value == "FAILED"
        assert workbook["START HERE"]["B8"].value > 0
    finally:
        workbook.close()


def test_formula_input_is_rejected_and_cli_writes_load_diagnostic(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    source = create_excel_template(tmp_path / "formula.xlsx", bundle=_economic_bundle())
    workbook = load_workbook(source, data_only=False, read_only=False)
    workbook["TRADES"]["H5"] = "=5+5"
    workbook.save(source)
    workbook.close()

    with pytest.raises(ExcelAdapterError, match=r"Formula not allowed.*tblTrades: H5"):
        load_excel_bundle(source)

    output_root = tmp_path / "results"
    assert (
        main(
            [
                "excel-build",
                "--workbook",
                str(source),
                "--output",
                str(output_root),
            ]
        )
        == 2
    )
    message = json.loads(capsys.readouterr().out)
    assert message["status"] == "FAILED"
    assert message["stage"] == "ExcelLoad"
    diagnostic = Path(message["workbook"])
    assert diagnostic.exists()
    assert not (output_root / "GTM_LATEST.xlsx").exists()
    assert _table_data_count(diagnostic, "VALIDATION", "tblValidation") == 1


def test_excel_template_cli_and_contract_errors(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    template = tmp_path / "template.xlsx"
    assert (
        main(
            [
                "excel-template",
                "--output",
                str(template),
                "--mapping",
                "docs/GTM_ACTIVE_SETUP_MAPPING_v0.3.csv",
            ]
        )
        == 0
    )
    created = json.loads(capsys.readouterr().out)
    assert created == {"status": "CREATED", "workbook": str(template.resolve())}

    wrong_suffix = tmp_path / "wrong.xlsm"
    wrong_suffix.write_bytes(b"not an Excel file")
    with pytest.raises(ExcelAdapterError, match="macro-free"):
        load_excel_bundle(wrong_suffix)

    workbook = load_workbook(template, data_only=False, read_only=False)
    workbook.remove(workbook["EVENT LEDGER"])
    workbook.save(template)
    workbook.close()
    with pytest.raises(ExcelAdapterError, match="EVENT LEDGER"):
        load_excel_bundle(template)

    assert (
        main(
            [
                "excel-template",
                "--output",
                str(tmp_path / "bad.xlsx"),
                "--mapping",
                str(tmp_path / "missing.csv"),
            ]
        )
        == 2
    )
    failed = json.loads(capsys.readouterr().out)
    assert failed["status"] == "FAILED"
    assert failed["stage"] == "ExcelTemplate"
