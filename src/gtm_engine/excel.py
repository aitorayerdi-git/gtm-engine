"""Macro-free Excel input/output adapter for the authoritative Python engine."""

from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from datetime import UTC, date, datetime
from decimal import Decimal
from enum import Enum
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from .io import BundleLoadError, publish_result
from .models import (
    ENGINE_VERSION,
    POLICY_VERSION,
    SCHEMA_VERSION,
    BookConfig,
    BuildConfig,
    BuildManifest,
    BuildResult,
    BuildStatus,
    CumulativePnlRow,
    CurvePrice,
    DeliveryElection,
    EventLedgerRow,
    ExposureRow,
    FixingPrice,
    FixingRow,
    FxRate,
    InitialExposure,
    InitialPnl,
    InputBundle,
    MarketCalendarDay,
    OperatingFlow,
    PnlRow,
    Severity,
    Trade,
    UnderlyingConfig,
    ValidationItem,
)
from .pipeline import build

TITLE_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
OUTPUT_HEADER_FILL = PatternFill("solid", fgColor="548235")
NOTE_FILL = PatternFill("solid", fgColor="D9EAF7")
ATTENTION_FILL = PatternFill("solid", fgColor="FFF2CC")
PASS_FILL = PatternFill("solid", fgColor="C6E0B4")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
WHITE_FONT = Font(name="Aptos", color="FFFFFF", bold=True)
BODY_FONT = Font(name="Aptos", size=10)
INPUT_FONT = Font(name="Aptos", size=10, color="0000FF")
TITLE_FONT = Font(name="Aptos Display", size=18, color="FFFFFF", bold=True)
THIN_GREY = Side(style="thin", color="D9E2F3")
TABLE_BORDER = Border(bottom=THIN_GREY)

DAILY_REPORT_SHEET = "Daily Report D2"
DAILY_REPORT_UNDERLYINGS = (
    "Brent Dated",
    "HH",
    "TTF DA",
    "TTF MA",
    "Index PVB",
    "Phys PVB",
    "TVB",
    "AVB",
    "PEG",
)
DAILY_REPORT_ZERO = Decimal("0.0000001")
DAILY_REPORT_FONT = Font(name="Arial", size=10)
DAILY_REPORT_BOLD_FONT = Font(name="Arial", size=10, bold=True)
DAILY_REPORT_NUMBER_FORMAT = "#,##0.0"

DATE_FIELDS = {
    "initial_market_date",
    "historical_start_date",
    "historical_end_date",
    "date",
    "trade_date",
    "decision_date",
    "start_date",
    "end_date",
    "market_date",
    "previous_market_date",
    "delivery_month",
    "curve_delivery_month",
    "delivery_day",
    "fixing_date",
    "applied_market_date",
    "price_lookup_date",
    "economic_date",
    "rate_date",
}
DATETIME_FIELDS = {"source_as_of", "started_at", "finished_at"}
VOLUME_FIELDS = {
    "daily_qty",
    "delivery_daily_qty",
    "exposure_volume",
    "fixing_volume",
    "signed_volume",
    "signed_volume_change",
}
PRICE_FIELDS = {"execution_price", "curve_price", "fixing_price"}
AMOUNT_FIELDS = {
    "amount",
    "fixing_amount",
    "exposure_mtm",
    "gross_delta_exposure_mtm",
    "trade_entry_adjustment",
    "delta_exposure_mtm",
    "logistics_source_amount",
    "logistical_costs",
    "fees_and_optimizations",
    "replication",
    "total_pnl",
    "initial_pnl",
    "daily_pnl",
    "cumulative_pnl",
}
RIGHT_ALIGNED_FIELDS = DATE_FIELDS | DATETIME_FIELDS | VOLUME_FIELDS | PRICE_FIELDS | AMOUNT_FIELDS


class ExcelAdapterError(BundleLoadError):
    """Workbook contract or content could not be loaded safely."""


class TableSpec:
    """Small immutable description of one workbook table."""

    __slots__ = ("is_input", "model", "note", "sheet", "table")

    def __init__(
        self,
        sheet: str,
        table: str,
        model: type[BaseModel],
        note: str,
        *,
        is_input: bool,
    ) -> None:
        self.sheet = sheet
        self.table = table
        self.model = model
        self.note = note
        self.is_input = is_input


INPUT_TABLES = (
    TableSpec("BOOKS", "tblBooks", BookConfig, "Authoritative active BOOK list.", is_input=True),
    TableSpec(
        "UNDERLYINGS",
        "tblUnderlyings",
        UnderlyingConfig,
        "Authoritative SETUP mapping and product methodology.",
        is_input=True,
    ),
    TableSpec(
        "MARKET CALENDAR",
        "tblMarketCalendar",
        MarketCalendarDay,
        "Configured Market Date axis. Do not infer holidays in the engine.",
        is_input=True,
    ),
    TableSpec(
        "INITIAL EXPOSURE",
        "tblInitialExposure",
        InitialExposure,
        "Open exposure at the close of the Initial Market Date.",
        is_input=True,
    ),
    TableSpec(
        "INITIAL PNL",
        "tblInitialPnl",
        InitialPnl,
        "Opening cumulative P&L bridge by BOOK; not first-day Daily P&L.",
        is_input=True,
    ),
    TableSpec("TRADES", "tblTrades", Trade, "Actual and simulation trade events.", is_input=True),
    TableSpec(
        "DELIVERY ELECTIONS",
        "tblDeliveryElections",
        DeliveryElection,
        "Daily BUY/SELL volume elected for physical delivery.",
        is_input=True,
    ),
    TableSpec(
        "CURVE PRICES",
        "tblCurvePrices",
        CurvePrice,
        "Daily curve marks by Market Date, Underlying and Delivery Month.",
        is_input=True,
    ),
    TableSpec(
        "FIXING PRICES",
        "tblFixingPrices",
        FixingPrice,
        "Fixing observations keyed by lookup date and configured price underlying.",
        is_input=True,
    ),
    TableSpec(
        "FX RATES",
        "tblFxRates",
        FxRate,
        "Daily currency units per EUR; prior available observation is used.",
        is_input=True,
    ),
    TableSpec(
        "OPERATING FLOWS",
        "tblOperatingFlows",
        OperatingFlow,
        "Daily Logistics, Fees/Optimizations and Replication flows by BOOK.",
        is_input=True,
    ),
)

OUTPUT_TABLES = (
    TableSpec(
        "VALIDATION",
        "tblValidation",
        ValidationItem,
        "Blocking errors, warnings and information generated by Python.",
        is_input=False,
    ),
    TableSpec("FIXINGS", "tblFixings", FixingRow, "Priced fixing rows.", is_input=False),
    TableSpec("EXPOSURE", "tblExposure", ExposureRow, "Daily open exposure.", is_input=False),
    TableSpec("DAILY PNL", "tblDailyPnl", PnlRow, "Daily P&L components.", is_input=False),
    TableSpec(
        "CUMULATIVE PNL",
        "tblCumulativePnl",
        CumulativePnlRow,
        "Opening bridge plus post-initial-date Daily P&L.",
        is_input=False,
    ),
    TableSpec(
        "EVENT LEDGER",
        "tblEventLedger",
        EventLedgerRow,
        "Auditable initial, trade and fixing events.",
        is_input=False,
    ),
)

CONTROL_FIELDS = tuple(BuildConfig.model_fields)
CONTROL_DESCRIPTIONS = {
    "model_id": "Stable model identifier.",
    "schema_version": "Input/output contract version; do not edit unless upgraded.",
    "policy_version": "Business-rule version; do not edit unless upgraded.",
    "engine_version": "Python engine version; do not edit unless upgraded.",
    "initial_market_date": "End-of-day opening state date (required).",
    "historical_start_date": "First requested Market Date (required).",
    "historical_end_date": "Last requested Market Date (required).",
    "simulation_status": "OFF excludes simulation trades; ON includes them.",
    "timezone": "Timezone used in the run manifest.",
    "logistics_sign": "Multiplier applied to Logistics source amounts.",
    "materiality": "Threshold below which exposure is treated as zero.",
}

SHEET_ORDER = (
    "START HERE",
    "CONTROL",
    *(spec.sheet for spec in INPUT_TABLES),
    *(spec.sheet for spec in OUTPUT_TABLES),
    "BUILD MANIFEST",
)


def _file_hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _display_header(field: str) -> str:
    return field.replace("_", " ").title()


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower().replace(" ", "_")


def _clean_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return None if stripped == "" else stripped
    return value


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.isoformat()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, sort_keys=True, default=str)
    return value


def _set_cell(cell: Any, value: Any) -> None:
    converted = _excel_value(value)
    cell.value = converted
    if isinstance(converted, str) and converted.startswith("="):
        cell.data_type = "s"


def _table_bounds(table: Table) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ExcelAdapterError(f"Excel table has an invalid range: {table.ref}")
    return min_col, min_row, max_col, max_row


def _find_table(workbook: Any, table_name: str) -> tuple[Worksheet, Table]:
    for sheet in workbook.worksheets:
        if table_name in sheet.tables:
            return sheet, sheet.tables[table_name]
    raise ExcelAdapterError(f"Required Excel table is missing: {table_name}")


def _require_contract(workbook: Any) -> None:
    missing_sheets = [
        name
        for name in SHEET_ORDER
        if name not in workbook.sheetnames and name not in {"FX RATES", "DELIVERY ELECTIONS"}
    ]
    if missing_sheets:
        raise ExcelAdapterError(f"Required worksheet(s) missing: {', '.join(missing_sheets)}")
    for table_name in (
        "tblControl",
        *(
            spec.table
            for spec in INPUT_TABLES
            if spec.table not in {"tblFxRates", "tblDeliveryElections"}
        ),
        "tblManifest",
    ):
        _find_table(workbook, table_name)
    for spec in OUTPUT_TABLES:
        _find_table(workbook, spec.table)


def _read_table(workbook: Any, spec: TableSpec) -> tuple[BaseModel, ...]:
    sheet, table = _find_table(workbook, spec.table)
    min_col, min_row, max_col, max_row = _table_bounds(table)
    cells = tuple(
        sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
    )
    if not cells:
        raise ExcelAdapterError(f"Excel table {spec.table} has no header row")
    headers = tuple(_normalize_header(cell.value) for cell in cells[0])
    expected = tuple(spec.model.model_fields)
    legacy_underlyings = spec.table == "tblUnderlyings" and headers == tuple(
        field for field in expected if field != "include_fixing_in_pnl"
    )
    if headers != expected and not legacy_underlyings:
        raise ExcelAdapterError(
            f"Excel table {spec.table} headers do not match the contract. "
            f"Expected: {', '.join(_display_header(name) for name in expected)}"
        )

    rows: list[BaseModel] = []
    for row_number, row_cells in enumerate(cells[1:], start=min_row + 1):
        values: list[Any] = []
        for cell in row_cells:
            if cell.data_type == "f":
                raise ExcelAdapterError(
                    "Formula not allowed in authoritative input table "
                    f"{spec.table}: {cell.coordinate}"
                )
            values.append(_clean_value(cell.value))
        if all(value is None for value in values):
            continue
        record = dict(zip(headers, values, strict=True))
        if legacy_underlyings:
            record["include_fixing_in_pnl"] = True
        source = record.get("source_row_id") or record.get("source_id")
        context = f", source={source}" if source else ""
        try:
            rows.append(spec.model.model_validate(record))
        except (ValidationError, ValueError) as exc:
            raise ExcelAdapterError(
                f"Cannot load {spec.table} at Excel row {row_number}{context}: {exc}"
            ) from exc
    return tuple(rows)


def _read_control(workbook: Any) -> BuildConfig:
    sheet, table = _find_table(workbook, "tblControl")
    min_col, min_row, max_col, max_row = _table_bounds(table)
    if max_col - min_col + 1 != 3:
        raise ExcelAdapterError("tblControl must contain Field, Value and Description columns")
    headers = tuple(
        _normalize_header(sheet.cell(min_row, column).value)
        for column in range(min_col, max_col + 1)
    )
    if headers != ("field", "value", "description"):
        raise ExcelAdapterError("tblControl headers must be Field, Value and Description")
    values: dict[str, Any] = {}
    for row_number in range(min_row + 1, max_row + 1):
        field_cell = sheet.cell(row_number, min_col)
        value_cell = sheet.cell(row_number, min_col + 1)
        if field_cell.data_type == "f" or value_cell.data_type == "f":
            coordinate = (
                field_cell.coordinate if field_cell.data_type == "f" else value_cell.coordinate
            )
            raise ExcelAdapterError(f"Formula not allowed in tblControl: {coordinate}")
        field = _normalize_header(field_cell.value)
        if not field:
            continue
        if field not in BuildConfig.model_fields:
            raise ExcelAdapterError(f"Unknown CONTROL field at Excel row {row_number}: {field}")
        if field in values:
            raise ExcelAdapterError(f"Duplicate CONTROL field at Excel row {row_number}: {field}")
        values[field] = _clean_value(value_cell.value)
    missing = [field for field in CONTROL_FIELDS if field not in values]
    if missing:
        raise ExcelAdapterError(f"CONTROL is missing field(s): {', '.join(missing)}")
    try:
        return BuildConfig.model_validate(values)
    except (ValidationError, ValueError) as exc:
        raise ExcelAdapterError(f"Cannot load CONTROL: {exc}") from exc


def load_excel_bundle(workbook_path: str | Path) -> InputBundle:
    """Read the strict input tables from a macro-free GTM interface workbook."""

    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise ExcelAdapterError(f"Excel workbook does not exist: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ExcelAdapterError("Excel interface must be a macro-free .xlsx workbook")
    try:
        workbook = load_workbook(path, data_only=False, keep_links=False)
    except Exception as exc:
        raise ExcelAdapterError(f"Cannot open Excel workbook: {exc}") from exc
    try:
        _require_contract(workbook)
        config = _read_control(workbook)
        loaded = {
            spec.table: (
                _read_table(workbook, spec)
                if spec.table not in {"tblFxRates", "tblDeliveryElections"}
                or spec.sheet in workbook.sheetnames
                else ()
            )
            for spec in INPUT_TABLES
        }
    finally:
        workbook.close()
    return InputBundle(
        config=config,
        books=loaded["tblBooks"],  # type: ignore[arg-type]
        underlyings=loaded["tblUnderlyings"],  # type: ignore[arg-type]
        calendar=loaded["tblMarketCalendar"],  # type: ignore[arg-type]
        initial_exposure=loaded["tblInitialExposure"],  # type: ignore[arg-type]
        initial_pnl=loaded["tblInitialPnl"],  # type: ignore[arg-type]
        trades=loaded["tblTrades"],  # type: ignore[arg-type]
        delivery_elections=loaded["tblDeliveryElections"],  # type: ignore[arg-type]
        curve_prices=loaded["tblCurvePrices"],  # type: ignore[arg-type]
        fixing_prices=loaded["tblFixingPrices"],  # type: ignore[arg-type]
        fx_rates=loaded["tblFxRates"],  # type: ignore[arg-type]
        operating_flows=loaded["tblOperatingFlows"],  # type: ignore[arg-type]
        input_hashes={"excel_workbook": _file_hash(path)},
    )


def load_setup_mapping(
    path: str | Path,
) -> tuple[tuple[BookConfig, ...], tuple[UnderlyingConfig, ...]]:
    """Load the reviewed active SETUP mapping used to seed a production template."""

    mapping_path = Path(path).expanduser().resolve()
    try:
        with mapping_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = tuple(csv.DictReader(handle))
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ExcelAdapterError(f"Cannot load SETUP mapping: {exc}") from exc
    books: list[BookConfig] = []
    underlyings: list[UnderlyingConfig] = []
    try:
        for row in rows:
            if row.get("record_type") == "BOOK":
                books.append(BookConfig(book=row["canonical_name"], active=True))
            elif row.get("record_type") == "UNDERLYING":
                underlyings.append(
                    UnderlyingConfig.model_validate(
                        {
                            "source_underlying": row["source_name"],
                            "canonical_underlying": row["canonical_name"],
                            "fixing_method": row["fixing_method"],
                            "unit": row["unit"],
                            "currency": "EUR",
                            "curve_underlying": row["curve_underlying"],
                            "fixing_price_underlying": row["fixing_price_underlying"],
                            "fixing_price_basis": row["fixing_price_basis"],
                            "include_fixing_in_pnl": row.get("include_fixing_in_pnl", "TRUE"),
                            "active": True,
                            "current_month_uses_next_curve": row["current_month_uses_next_curve"],
                        }
                    )
                )
    except (KeyError, ValidationError, ValueError) as exc:
        raise ExcelAdapterError(f"Invalid SETUP mapping row: {exc}") from exc
    return tuple(books), tuple(underlyings)


def _title_and_note(sheet: Worksheet, title: str, note: str, width: int) -> None:
    final_column = get_column_letter(max(width, 2))
    sheet.merge_cells(f"A1:{final_column}1")
    sheet["A1"] = title
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(f"A2:{final_column}2")
    sheet["A2"] = note
    sheet["A2"].fill = NOTE_FILL
    sheet["A2"].font = BODY_FONT
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32
    sheet.sheet_view.showGridLines = False


def _column_format(field: str) -> str:
    if field in DATE_FIELDS:
        return "yyyy-mm-dd"
    if field in DATETIME_FIELDS:
        return "yyyy-mm-dd hh:mm:ss"
    if field in VOLUME_FIELDS:
        return "#,##0.000000;[Red](#,##0.000000);-"
    if field in PRICE_FIELDS:
        return "#,##0.00000000;[Red](#,##0.00000000);-"
    if field in AMOUNT_FIELDS:
        return "#,##0.00;[Red](#,##0.00);-"
    if field in {"logistics_sign", "materiality"}:
        return "0.00000000;[Red](0.00000000);-"
    if field in {"elapsed_seconds"}:
        return "0.000"
    if field in {"peak_memory_bytes"}:
        return "#,##0"
    return "General"


def _column_width(field: str, rows: Sequence[Mapping[str, Any]]) -> float:
    header = _display_header(field)
    sample = max((len(str(row.get(field, ""))) for row in rows[:200]), default=0)
    if field in {"message", "remediation", "comment"}:
        return 42
    if field in {"economic_key", "source_row_id", "source_id", "event_id"}:
        return min(max(len(header) + 2, sample + 2, 18), 34)
    if field in DATE_FIELDS or field in DATETIME_FIELDS:
        return 19 if field in DATETIME_FIELDS else 13
    return min(max(len(header) + 2, sample + 2, 11), 28)


def _records(rows: Iterable[BaseModel]) -> list[dict[str, Any]]:
    return [row.model_dump(mode="python") for row in rows]


def _add_table_sheet(
    workbook: Workbook,
    spec: TableSpec,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    sheet = workbook.create_sheet(spec.sheet)
    fields = tuple(spec.model.model_fields)
    _title_and_note(sheet, spec.sheet, spec.note, len(fields))
    header_row = 4
    data_rows = list(rows) or [{}]
    for column, field in enumerate(fields, start=1):
        header = sheet.cell(header_row, column, _display_header(field))
        header.fill = INPUT_HEADER_FILL if spec.is_input else OUTPUT_HEADER_FILL
        header.font = WHITE_FONT
        header.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.column_dimensions[get_column_letter(column)].width = _column_width(field, data_rows)
        for offset, record in enumerate(data_rows, start=1):
            cell = sheet.cell(header_row + offset, column)
            _set_cell(cell, record.get(field))
            cell.font = INPUT_FONT if spec.is_input else BODY_FONT
            cell.border = TABLE_BORDER
            cell.number_format = _column_format(field)
            cell.alignment = Alignment(
                horizontal="right" if field in RIGHT_ALIGNED_FIELDS else "left",
                vertical="top",
                wrap_text=field in {"message", "remediation", "comment", "economic_key"},
            )
    last_row = header_row + len(data_rows)
    table = Table(
        displayName=spec.table,
        ref=f"A{header_row}:{get_column_letter(len(fields))}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2" if spec.is_input else "TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A5"
    sheet.sheet_properties.tabColor = "4472C4" if spec.is_input else "70AD47"


def _control_rows(config: BuildConfig | None) -> list[dict[str, Any]]:
    default_values: dict[str, Any] = {
        "model_id": "GTM",
        "schema_version": SCHEMA_VERSION,
        "policy_version": POLICY_VERSION,
        "engine_version": ENGINE_VERSION,
        "initial_market_date": None,
        "historical_start_date": None,
        "historical_end_date": None,
        "simulation_status": "OFF",
        "timezone": "Europe/Madrid",
        "logistics_sign": Decimal("-1"),
        "materiality": Decimal("0.0000001"),
    }
    if config is not None:
        default_values.update(config.model_dump(mode="python"))
    return [
        {
            "field": _display_header(field),
            "value": default_values[field],
            "description": CONTROL_DESCRIPTIONS[field],
        }
        for field in CONTROL_FIELDS
    ]


def _add_control_sheet(workbook: Workbook, config: BuildConfig | None) -> None:
    sheet = workbook.create_sheet("CONTROL")
    _title_and_note(
        sheet,
        "CONTROL",
        "Blue values are user inputs. Yellow cells are required before a build.",
        3,
    )
    headers = ("Field", "Value", "Description")
    for column, header_text in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header_text)
        cell.fill = INPUT_HEADER_FILL
        cell.font = WHITE_FONT
    rows = _control_rows(config)
    for row_number, record in enumerate(rows, start=5):
        for column, field in enumerate(("field", "value", "description"), start=1):
            cell = sheet.cell(row_number, column)
            _set_cell(cell, record[field])
            cell.font = INPUT_FONT if field == "value" else BODY_FONT
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(wrap_text=field == "description", vertical="top")
        field_name = CONTROL_FIELDS[row_number - 5]
        sheet.cell(row_number, 2).number_format = _column_format(field_name)
        if field_name in {"initial_market_date", "historical_start_date", "historical_end_date"}:
            sheet.cell(row_number, 2).fill = ATTENTION_FILL
            sheet.cell(row_number, 2).comment = Comment(
                "Required control. Enter a real Excel date.", "User"
            )
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 60
    sheet.row_dimensions[2].height = 28
    table = Table(displayName="tblControl", ref=f"A4:C{4 + len(rows)}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    simulation_row = 5 + CONTROL_FIELDS.index("simulation_status")
    validation = DataValidation(type="list", formula1='"OFF,ON"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(sheet.cell(simulation_row, 2))
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "17365D"


def _add_start_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("START HERE")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "17365D"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "GTM — Excel Interface / Python Engine"
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:H2")
    sheet["A2"] = (
        "Macro-free interface. Excel holds inputs and displays outputs; Python is the only "
        "calculation authority."
    )
    sheet["A2"].fill = NOTE_FILL
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 30

    labels = (
        ("A4", "BUILD STATUS", "B4", "NOT BUILT"),
        ("A5", "Build ID", "B5", "—"),
        ("A6", "Run ID", "B6", "—"),
        ("A7", "Finished", "B7", "—"),
        ("A8", "Errors", "B8", 0),
        ("A9", "Warnings", "B9", 0),
    )
    for label_cell, label, value_cell, value in labels:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(name="Aptos", bold=True)
        sheet[value_cell] = value
        sheet[value_cell].border = TABLE_BORDER
    sheet["B4"].fill = ATTENTION_FILL
    sheet["B4"].font = Font(name="Aptos", bold=True)

    sheet.merge_cells("A11:H11")
    sheet["A11"] = "DAILY WORKFLOW"
    sheet["A11"].fill = SECTION_FILL
    sheet["A11"].font = WHITE_FONT
    steps = (
        "1. Update the blue input tables. Never put formulas in input tables.",
        "2. Save this workbook and close it before building.",
        "3. Run: .venv/bin/gtm-engine excel-build --workbook <this.xlsx> "
        "--output outputs/gtm_excel_runs",
        "4. Open outputs/gtm_excel_runs/GTM_LATEST.xlsx only if status is PUBLISHED.",
        "5. If a build fails, open the retained GTM_Failed.xlsx and read VALIDATION.",
    )
    for row_number, instruction in enumerate(steps, start=12):
        sheet.cell(row_number, 1, instruction)
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=8)
        sheet.cell(row_number, 1).alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row_number].height = 24

    sheet.merge_cells("A18:H18")
    sheet["A18"] = "CONTRACT AND COLOUR LEGEND"
    sheet["A18"].fill = SECTION_FILL
    sheet["A18"].font = WHITE_FONT
    legend = (
        ("Blue text", "Editable user input", "0000FF", None),
        ("Yellow fill", "Required control needs attention", "000000", ATTENTION_FILL),
        ("Green tab/header", "Python-generated output", "000000", PASS_FILL),
        (
            "Red status",
            "Blocking failure; previous latest result remains unchanged",
            "000000",
            FAIL_FILL,
        ),
    )
    for row_number, (label, meaning, color, fill) in enumerate(legend, start=19):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 1).font = Font(name="Aptos", color=color, bold=True)
        if fill is not None:
            sheet.cell(row_number, 1).fill = fill
        sheet.cell(row_number, 2, meaning)
        sheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=8)
    sheet.column_dimensions["A"].width = 24
    for column in "BCDEFGH":
        sheet.column_dimensions[column].width = 16
    sheet.freeze_panes = "A4"


def _add_manifest_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("BUILD MANIFEST")
    _title_and_note(
        sheet,
        "BUILD MANIFEST",
        "Immutable run identity, hashes, row counts, totals and runtime evidence from Python.",
        2,
    )
    sheet["A4"] = "Field"
    sheet["B4"] = "Value"
    for cell in sheet[4]:
        cell.fill = OUTPUT_HEADER_FILL
        cell.font = WHITE_FONT
    sheet["A5"] = None
    sheet["B5"] = None
    table = Table(displayName="tblManifest", ref="A4:B5")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 80
    sheet.freeze_panes = "A5"
    sheet.sheet_properties.tabColor = "70AD47"


def _apply_input_validations(workbook: Workbook) -> None:
    validations: dict[tuple[str, str], tuple[str, ...]] = {
        ("UNDERLYINGS", "fixing_method"): (
            "WITHINDAY",
            "DAY_AHEAD",
            "HEREN",
            "MONTH_AHEAD",
            "BRENT_HH",
        ),
        ("UNDERLYINGS", "fixing_price_basis"): ("FIXING_DATE", "DELIVERY_DAY"),
        ("TRADES", "side"): ("BUY", "SELL"),
        ("TRADES", "trade_source"): ("ACTUAL", "SIMULATION"),
    }
    spec_by_sheet = {spec.sheet: spec for spec in INPUT_TABLES}
    for (sheet_name, field), choices in validations.items():
        sheet = workbook[sheet_name]
        fields = tuple(spec_by_sheet[sheet_name].model.model_fields)
        column = get_column_letter(fields.index(field) + 1)
        validation = DataValidation(
            type="list", formula1=f'"{",".join(choices)}"', allow_blank=False
        )
        sheet.add_data_validation(validation)
        validation.add(f"{column}5:{column}5000")

    for sheet_name, field in (
        ("BOOKS", "active"),
        ("UNDERLYINGS", "active"),
        ("UNDERLYINGS", "current_month_uses_next_curve"),
        ("MARKET CALENDAR", "is_market_day"),
    ):
        sheet = workbook[sheet_name]
        fields = tuple(spec_by_sheet[sheet_name].model.model_fields)
        column = get_column_letter(fields.index(field) + 1)
        validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        sheet.add_data_validation(validation)
        validation.add(f"{column}5:{column}5000")

    trades = workbook["TRADES"]
    qty_column = get_column_letter(tuple(Trade.model_fields).index("daily_qty") + 1)
    trades.conditional_formatting.add(
        f"{qty_column}5:{qty_column}5000",
        CellIsRule(operator="lessThan", formula=["0"], fill=FAIL_FILL),  # type: ignore[no-untyped-call]
    )


def create_excel_template(
    destination: str | Path,
    *,
    bundle: InputBundle | None = None,
    books: Sequence[BookConfig] = (),
    underlyings: Sequence[UnderlyingConfig] = (),
) -> Path:
    """Create a clean, macro-free GTM workbook; optionally populate a test/input bundle."""

    path = Path(destination).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is None:
        raise ExcelAdapterError("Cannot initialize Excel workbook")
    workbook.remove(default_sheet)
    _add_start_sheet(workbook)
    _add_control_sheet(workbook, bundle.config if bundle else None)
    bundle_rows: dict[str, Sequence[BaseModel]] = {}
    if bundle is not None:
        bundle_rows = {
            "tblBooks": bundle.books,
            "tblUnderlyings": bundle.underlyings,
            "tblMarketCalendar": bundle.calendar,
            "tblInitialExposure": bundle.initial_exposure,
            "tblInitialPnl": bundle.initial_pnl,
            "tblTrades": bundle.trades,
            "tblDeliveryElections": bundle.delivery_elections,
            "tblCurvePrices": bundle.curve_prices,
            "tblFixingPrices": bundle.fixing_prices,
            "tblFxRates": bundle.fx_rates,
            "tblOperatingFlows": bundle.operating_flows,
        }
    else:
        bundle_rows = {"tblBooks": books, "tblUnderlyings": underlyings}
    for spec in INPUT_TABLES:
        _add_table_sheet(workbook, spec, _records(bundle_rows.get(spec.table, ())))
    for spec in OUTPUT_TABLES:
        _add_table_sheet(workbook, spec, [])
    _add_manifest_sheet(workbook)
    _apply_input_validations(workbook)
    workbook["START HERE"].sheet_view.zoomScale = 95
    workbook.active = 0
    workbook.properties.title = "GTM Excel Interface / Python Engine"
    workbook.properties.subject = "Macro-free gas-trading model input and output contract"
    workbook.properties.creator = "GTM Python Engine"
    _atomic_save(workbook, path)
    workbook.close()
    return path


def _replace_table_records(
    workbook: Any,
    spec: TableSpec,
    rows: Sequence[BaseModel],
) -> None:
    sheet, table = _find_table(workbook, spec.table)
    fields = tuple(spec.model.model_fields)
    min_col, min_row, max_col, old_max_row = _table_bounds(table)
    if max_col - min_col + 1 != len(fields):
        raise ExcelAdapterError(f"Output table {spec.table} has an invalid column count")
    records = _records(rows)
    new_data_count = max(1, len(records))
    new_max_row = min_row + new_data_count
    clear_max_row = max(old_max_row, new_max_row)
    for row in sheet.iter_rows(
        min_row=min_row + 1,
        max_row=clear_max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None
    for row_offset, record in enumerate(records, start=1):
        for column_offset, field in enumerate(fields):
            cell = sheet.cell(min_row + row_offset, min_col + column_offset)
            _set_cell(cell, record.get(field))
            cell.font = BODY_FONT
            cell.border = TABLE_BORDER
            cell.number_format = _column_format(field)
            cell.alignment = Alignment(
                horizontal="right" if field in RIGHT_ALIGNED_FIELDS else "left",
                vertical="top",
                wrap_text=field in {"message", "remediation", "comment", "economic_key"},
            )
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"


def _replace_manifest(workbook: Any, manifest: BuildManifest) -> None:
    sheet, table = _find_table(workbook, "tblManifest")
    min_col, min_row, max_col, old_max_row = _table_bounds(table)
    records = manifest.model_dump(mode="python")
    rows = [(field, records[field]) for field in BuildManifest.model_fields]
    new_max_row = min_row + len(rows)
    for row in sheet.iter_rows(
        min_row=min_row + 1,
        max_row=max(old_max_row, new_max_row),
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None
    for row_offset, (field, value) in enumerate(rows, start=1):
        field_cell = sheet.cell(min_row + row_offset, min_col)
        value_cell = sheet.cell(min_row + row_offset, min_col + 1)
        field_cell.value = _display_header(field)
        _set_cell(value_cell, value)
        field_cell.font = BODY_FONT
        value_cell.font = BODY_FONT
        field_cell.border = TABLE_BORDER
        value_cell.border = TABLE_BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        value_cell.number_format = _column_format(field)
    table.ref = f"A{min_row}:B{new_max_row}"


def _update_start_sheet(workbook: Any, result: BuildResult) -> None:
    sheet = workbook["START HERE"]
    counts = result.manifest.validation_counts
    status = result.manifest.status.value
    values = {
        "B4": status,
        "B5": result.manifest.build_id,
        "B6": result.manifest.run_id,
        "B7": result.manifest.finished_at.isoformat(),
        "B8": counts.get("ERROR", 0),
        "B9": counts.get("WARNING", 0),
    }
    for coordinate, value in values.items():
        _set_cell(sheet[coordinate], value)
    sheet["B4"].fill = PASS_FILL if result.manifest.status is BuildStatus.PUBLISHED else FAIL_FILL
    sheet["B4"].font = Font(name="Aptos", bold=True)


def _next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def _daily_report_period(result: BuildResult) -> tuple[date, date]:
    if not result.cumulative_pnl:
        raise ExcelAdapterError("Published result has no Cumulative P&L dates for Daily Report D2")
    d2 = max(row.market_date for row in result.cumulative_pnl)
    previous_dates = {
        row.previous_market_date for row in result.cumulative_pnl if row.market_date == d2
    }
    if len(previous_dates) != 1:
        raise ExcelAdapterError("Daily Report D2 has an ambiguous Previous Market Date")
    return previous_dates.pop(), d2


def _daily_report_book_rows(rows: Sequence[PnlRow]) -> list[tuple[Any, ...]]:
    totals: defaultdict[str, list[Decimal]] = defaultdict(lambda: [Decimal("0") for _ in range(6)])
    for row in rows:
        values = (
            row.total_pnl,
            row.delta_exposure_mtm,
            row.fixing_amount,
            row.logistical_costs,
            row.fees_and_optimizations,
            row.replication,
        )
        for index, value in enumerate(values):
            totals[row.book][index] += value
    return [
        (book, *values)
        for book, values in sorted(totals.items())
        if any(abs(value) > DAILY_REPORT_ZERO for value in values)
    ]


def _daily_report_book_month_rows(rows: Sequence[PnlRow]) -> list[tuple[Any, ...]]:
    totals: defaultdict[tuple[str, date], list[Decimal]] = defaultdict(
        lambda: [Decimal("0"), Decimal("0")]
    )
    for row in rows:
        if row.delivery_month is None:
            continue
        key = (row.book, row.delivery_month)
        totals[key][0] += row.delta_exposure_mtm
        totals[key][1] += row.fixing_amount
    return [
        (book, delivery_month, *values)
        for (book, delivery_month), values in sorted(totals.items())
        if any(abs(value) > DAILY_REPORT_ZERO for value in values)
    ]


def _daily_report_exposure_rows(
    d2: date,
    rows: Sequence[PnlRow],
) -> list[tuple[Any, ...]]:
    totals: defaultdict[tuple[date, str], Decimal] = defaultdict(lambda: Decimal("0"))
    delivery_months: set[date] = set()
    for row in rows:
        if row.delivery_month is None or row.underlying not in DAILY_REPORT_UNDERLYINGS:
            continue
        delivery_months.add(row.delivery_month)
        totals[(row.delivery_month, row.underlying)] += row.delta_exposure_mtm

    start = date(d2.year, 1, 1)
    end = date(d2.year + 2, 12, 1)
    if delivery_months:
        start = min(start, min(delivery_months))
        end = max(end, max(delivery_months))

    output: list[tuple[Any, ...]] = []
    delivery_month = start
    while delivery_month <= end:
        output.append(
            (
                delivery_month,
                *(totals[(delivery_month, underlying)] for underlying in DAILY_REPORT_UNDERLYINGS),
            )
        )
        delivery_month = _next_month(delivery_month)
    return output


def _write_daily_report_rows(
    sheet: Worksheet,
    start_row: int,
    rows: Sequence[tuple[Any, ...]],
    *,
    month_column: int | None = None,
) -> None:
    for row_offset, values in enumerate(rows):
        row_number = start_row + row_offset
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column)
            _set_cell(cell, value)
            cell.font = DAILY_REPORT_FONT
            cell.alignment = Alignment(horizontal="right" if column == 1 else "center")
            if month_column == column:
                cell.number_format = "mmm/yy"
            elif isinstance(value, (Decimal, float, int)):
                cell.number_format = DAILY_REPORT_NUMBER_FORMAT


def _write_daily_report_d2(workbook: Any, result: BuildResult) -> None:
    if DAILY_REPORT_SHEET in workbook.sheetnames:
        workbook.remove(workbook[DAILY_REPORT_SHEET])
    if result.manifest.status is not BuildStatus.PUBLISHED:
        return

    d1, d2 = _daily_report_period(result)
    pnl_rows = tuple(row for row in result.pnl if row.market_date == d2)
    book_rows = _daily_report_book_rows(pnl_rows)
    book_month_rows = _daily_report_book_month_rows(pnl_rows)
    exposure_rows = _daily_report_exposure_rows(d2, pnl_rows)

    position = workbook.sheetnames.index("EVENT LEDGER")
    sheet = workbook.create_sheet(DAILY_REPORT_SHEET, position)
    sheet.sheet_properties.tabColor = "70AD47"
    sheet.sheet_view.showGridLines = True

    widths = {
        "A": 102.26953125,
        "B": 15,
        "C": 24,
        "D": 24.54296875,
        "E": 25.26953125,
        "F": 32.81640625,
        "G": 21.54296875,
        "H": 9.7265625,
        "I": 4.7265625,
        "J": 4.81640625,
        "K": 13,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet["A1"] = DAILY_REPORT_SHEET
    sheet["A1"].font = DAILY_REPORT_BOLD_FONT
    sheet["B1"] = "D1"
    sheet["C1"] = d1
    sheet["D1"] = "D2"
    sheet["E1"] = d2
    for coordinate in ("B1", "C1", "D1", "E1"):
        sheet[coordinate].font = DAILY_REPORT_FONT
        sheet[coordinate].alignment = Alignment(horizontal="center")
    sheet["C1"].number_format = "dd/mm/yy"
    sheet["E1"].number_format = "dd/mm/yy"
    sheet.row_dimensions[1].height = 13

    sheet["A3"] = "1) Delta PnL D2 vs D1 by BOOK"
    sheet["A3"].font = DAILY_REPORT_BOLD_FONT
    section_one_headers = (
        "BOOK",
        "Total Delta PnL",
        "Sum Delta Exposure MtM",
        "Sum Delta Fixing Amount",
        "Sum Delta Logistical Costs",
        "Sum Delta Fees and Optimizations",
        "Sum Delta Replication",
    )
    for column, header in enumerate(section_one_headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = DAILY_REPORT_BOLD_FONT
        cell.alignment = Alignment(horizontal="right" if column == 1 else "center")
    _write_daily_report_rows(sheet, 5, book_rows)

    section_two_title_row = max(16, 5 + len(book_rows) + 2)
    sheet.cell(section_two_title_row, 1, "2) Delta PnL D2 vs D1 by BOOK and Delivery Month")
    sheet.cell(section_two_title_row, 1).font = DAILY_REPORT_BOLD_FONT
    section_two_header_row = section_two_title_row + 1
    section_two_headers = (
        "BOOK",
        "Delivery Month",
        "Sum Delta Exposure MtM",
        "Sum Delta Fixing Amount",
    )
    for column, header in enumerate(section_two_headers, start=1):
        cell = sheet.cell(section_two_header_row, column, header)
        cell.font = DAILY_REPORT_BOLD_FONT
        cell.alignment = Alignment(horizontal="right" if column == 1 else "center")
    section_two_data_row = section_two_header_row + 1
    _write_daily_report_rows(
        sheet,
        section_two_data_row,
        book_month_rows,
        month_column=2,
    )

    section_three_title_row = section_two_data_row + len(book_month_rows) + 2
    sheet.cell(
        section_three_title_row,
        1,
        "3) Delta Exposure D2 vs D1 by BOOK, Delivery Month and Underlying",
    )
    sheet.cell(section_three_title_row, 1).font = DAILY_REPORT_BOLD_FONT
    sheet.cell(
        section_three_title_row + 1,
        1,
        "Note: section 3 shows selected underlyings only. Values are Delta Exposure MtM "
        "from the Python engine. Zeros are shown explicitly.",
    )
    sheet.cell(section_three_title_row + 1, 1).font = DAILY_REPORT_FONT
    sheet.cell(section_three_title_row + 3, 1, "ALL BOOKS")
    sheet.cell(section_three_title_row + 3, 1).font = DAILY_REPORT_BOLD_FONT

    section_three_header_row = section_three_title_row + 4
    section_three_headers = ("Delivery Month", *DAILY_REPORT_UNDERLYINGS)
    for column, header in enumerate(section_three_headers, start=2):
        cell = sheet.cell(section_three_header_row, column, header)
        cell.font = DAILY_REPORT_BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
    _write_daily_report_rows(
        sheet,
        section_three_header_row + 1,
        tuple((None, *row) for row in exposure_rows),
        month_column=2,
    )

    for row_number in (
        3,
        4,
        section_two_title_row,
        section_two_header_row,
        section_three_title_row,
        section_three_title_row + 3,
        section_three_header_row,
    ):
        sheet.row_dimensions[row_number].height = 13


def _atomic_save(workbook: Any, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        dir=destination.parent,
        prefix=f".{destination.stem}-",
        suffix=".xlsx",
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
        destination.chmod(0o644)
    finally:
        if temporary.exists():
            temporary.unlink()


def _atomic_copy(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        dir=destination.parent,
        prefix=f".{destination.stem}-",
        suffix=destination.suffix,
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
    try:
        shutil.copyfile(source, temporary)
        os.replace(temporary, destination)
        destination.chmod(0o644)
    finally:
        if temporary.exists():
            temporary.unlink()


def write_result_workbook(
    source_workbook: str | Path,
    destination: str | Path,
    result: BuildResult,
) -> Path:
    """Copy the source interface and replace only Python-owned result tables."""

    source = Path(source_workbook).expanduser().resolve()
    target = Path(destination).expanduser().resolve()
    try:
        workbook = load_workbook(source, data_only=False, keep_links=False)
    except Exception as exc:
        raise ExcelAdapterError(f"Cannot reopen source workbook for output: {exc}") from exc
    try:
        _require_contract(workbook)
        output_rows: dict[str, Sequence[BaseModel]] = {
            "tblValidation": result.validation,
            "tblFixings": result.fixings,
            "tblExposure": result.exposure,
            "tblDailyPnl": result.pnl,
            "tblCumulativePnl": result.cumulative_pnl,
            "tblEventLedger": result.event_ledger,
        }
        for spec in OUTPUT_TABLES:
            _replace_table_records(workbook, spec, output_rows[spec.table])
        _write_daily_report_d2(workbook, result)
        _replace_manifest(workbook, result.manifest)
        _update_start_sheet(workbook, result)
        workbook.active = 0
        _atomic_save(workbook, target)
    finally:
        workbook.close()
    return target


def build_excel_workbook(
    workbook_path: str | Path,
    output_root: str | Path,
) -> tuple[BuildResult, Path, Path]:
    """Load Excel, run Python, publish normalized evidence, and return results to Excel."""

    source = Path(workbook_path).expanduser().resolve()
    source_hash = _file_hash(source)
    bundle = load_excel_bundle(source)
    result = build(bundle)
    published, run_directory = publish_result(result, output_root)
    result_name = (
        "GTM_Result.xlsx"
        if published.manifest.status is BuildStatus.PUBLISHED
        else "GTM_Failed.xlsx"
    )
    result_workbook = write_result_workbook(source, run_directory / result_name, published)
    if _file_hash(source) != source_hash:
        raise ExcelAdapterError("Source workbook changed during the build; result was not promoted")
    if published.manifest.status is BuildStatus.PUBLISHED:
        _atomic_copy(result_workbook, Path(output_root).expanduser().resolve() / "GTM_LATEST.xlsx")
    return published, run_directory, result_workbook


def publish_excel_load_failure(
    source_workbook: str | Path,
    output_root: str | Path,
    message: str,
) -> Path | None:
    """Best-effort diagnostic workbook for a contract/load error; never promotes latest."""

    source = Path(source_workbook).expanduser().resolve()
    if not source.exists() or source.suffix.lower() != ".xlsx":
        return None
    run_name = f"load-{datetime.now(UTC).strftime('%Y%m%dT%H%M%S%fZ')}-{_file_hash(source)[:8]}"
    run_directory = Path(output_root).expanduser().resolve() / "failed" / run_name
    run_directory.mkdir(parents=True, exist_ok=False)
    try:
        workbook = load_workbook(source, data_only=False, keep_links=False)
        _require_contract(workbook)
        validation = ValidationItem(
            build_id="LOAD-FAILED",
            stage="Load",
            severity=Severity.ERROR,
            code="EXCEL_LOAD_ERROR",
            message=message,
            remediation="Correct the named workbook table/cell, save, and run again.",
        )
        for spec in OUTPUT_TABLES:
            rows: Sequence[BaseModel] = (validation,) if spec.table == "tblValidation" else ()
            _replace_table_records(workbook, spec, rows)
        sheet = workbook["START HERE"]
        sheet["B4"] = "FAILED"
        sheet["B4"].fill = FAIL_FILL
        sheet["B5"] = "LOAD-FAILED"
        sheet["B6"] = run_name
        sheet["B7"] = datetime.now(UTC).isoformat()
        sheet["B8"] = 1
        sheet["B9"] = 0
        target = run_directory / "GTM_Failed.xlsx"
        _atomic_save(workbook, target)
        workbook.close()
        return target
    except Exception:
        shutil.rmtree(run_directory, ignore_errors=True)
        return None


def workbook_formula_cells(path: str | Path) -> tuple[str, ...]:
    """Return every formula coordinate for verification; GTM interface files should have none."""

    workbook = load_workbook(Path(path).expanduser().resolve(), data_only=False, read_only=False)
    try:
        return tuple(
            f"{sheet.title}!{cell.coordinate}"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        )
    finally:
        workbook.close()
