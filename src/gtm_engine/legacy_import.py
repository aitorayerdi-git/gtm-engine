"""Read-only migration from the legacy GTM workbook into the v0.3 contract."""

from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import ValidationError

from .excel import create_excel_template
from .io import write_bundle
from .models import (
    ENGINE_VERSION,
    BookConfig,
    BuildConfig,
    CurvePrice,
    FixingMethod,
    FixingPrice,
    InitialExposure,
    InitialPnl,
    InputBundle,
    MarketCalendarDay,
    OperatingFlow,
    PriceDateBasis,
    Severity,
    Side,
    SimulationStatus,
    Trade,
    TradeSource,
    UnderlyingConfig,
)

IMPORTER_VERSION = "0.3.0"
REQUIRED_SHEETS = (
    "SETUP",
    "CALENDAR",
    "TRADES",
    "SIMULATION TRADES",
    "INITIAL POSITION",
    "INITIAL POSITION DATA",
    "PROCESS",
    "FIXING PRICES",
    "COSTS",
    "Foto FO",
)

METHODS = {
    "withinday": FixingMethod.WITHINDAY,
    "day ahead": FixingMethod.DAY_AHEAD,
    "metodología heren": FixingMethod.HEREN,
    "month ahead": FixingMethod.MONTH_AHEAD,
    "brent & hh": FixingMethod.BRENT_HH,
}

SOURCE_RANGES = {
    "books": "SETUP!B3:C15",
    "underlyings_main": "SETUP!G3:J11",
    "underlyings_pvb": "SETUP!M3:P11",
    "calendar": "CALENDAR!A4:E1100",
    "initial_exposure": "INITIAL POSITION DATA!A2:E8209",
    "initial_pnl": "INITIAL POSITION!A5:C17",
    "trades_actual": "TRADES!A3:I478",
    "trades_simulation": "SIMULATION TRADES!A3:O2268",
    "curves_ttf": "TTF!D9:BF601",
    "curves_ttf_prompt": "TTF!AS7:BA601",
    "curves_brent": "Brent Dated!D9:AX588",
    "curves_hh": "HH!D9:AW874",
    "curves_pvb_spread": "PVB-TTF!B3:BY412",
    "curves_peg_spread": "PEG-TTF!B3:CA408",
    "fixing_prices": "FIXING PRICES!A6:S1102",
    "logistics": "COSTS!A5:N504",
    "fees": "COSTS!A5:P504",
    "optimizations": "Foto FO!B6:S1118",
    "replication": "Foto FO!B6:V1118",
}


class LegacyImportError(ValueError):
    """The source workbook could not be converted safely."""


@dataclass(frozen=True)
class ImportIssue:
    severity: Severity
    code: str
    message: str
    source: str | None = None
    count: int | None = None
    source_rows: tuple[str, ...] = ()

    def as_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["severity"] = self.severity.value
        result["source_rows"] = list(self.source_rows)
        return result


@dataclass
class LegacyImportReport:
    source_path: str
    source_sha256: str
    source_modified_at: str
    imported_at: str
    initial_market_date: str = ""
    historical_start_date: str = ""
    historical_end_date: str = ""
    simulation_status: str = ""
    extracted_row_counts: dict[str, int] = field(default_factory=dict)
    skipped_row_counts: dict[str, int] = field(default_factory=dict)
    source_ranges: dict[str, str] = field(default_factory=lambda: dict(SOURCE_RANGES))
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return sum(issue.severity is Severity.ERROR for issue in self.issues)

    @property
    def warning_count(self) -> int:
        return sum(issue.severity is Severity.WARNING for issue in self.issues)

    @property
    def status(self) -> str:
        return "CREATED_WITH_REVIEW_ITEMS" if self.error_count else "CREATED"

    def add(
        self,
        severity: Severity,
        code: str,
        message: str,
        *,
        source: str | None = None,
        count: int | None = None,
        source_rows: tuple[str, ...] = (),
    ) -> None:
        self.issues.append(
            ImportIssue(
                severity=severity,
                code=code,
                message=message,
                source=source,
                count=count,
                source_rows=source_rows,
            )
        )

    def as_dict(self) -> dict[str, Any]:
        return {
            "importer_version": IMPORTER_VERSION,
            "engine_version": ENGINE_VERSION,
            "status": self.status,
            "source": {
                "path": self.source_path,
                "sha256": self.source_sha256,
                "modified_at": self.source_modified_at,
                "unchanged_during_import": True,
            },
            "controls": {
                "initial_market_date": self.initial_market_date,
                "historical_start_date": self.historical_start_date,
                "historical_end_date": self.historical_end_date,
                "simulation_status": self.simulation_status,
            },
            "extracted_row_counts": self.extracted_row_counts,
            "skipped_row_counts": self.skipped_row_counts,
            "source_ranges": self.source_ranges,
            "issue_counts": {
                "ERROR": self.error_count,
                "WARNING": self.warning_count,
                "INFO": sum(issue.severity is Severity.INFO for issue in self.issues),
            },
            "issues": [issue.as_dict() for issue in self.issues],
        }


@dataclass(frozen=True)
class LegacyImportResult:
    bundle: InputBundle
    report: LegacyImportReport
    output_directory: Path
    normalized_bundle: Path
    excel_workbook: Path
    audit_json: Path
    issues_csv: Path


def _file_hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _date_value(value: Any, source: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                pass
    raise LegacyImportError(f"Expected a date at {source}; found {value!r}")


def _optional_date(value: Any) -> date | None:
    try:
        return _date_value(value, "optional date")
    except LegacyImportError:
        return None


def _decimal_value(value: Any, source: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise LegacyImportError(f"Expected a number at {source}; found {value!r}")
    try:
        result = value if isinstance(value, Decimal) else Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise LegacyImportError(f"Expected a number at {source}; found {value!r}") from exc
    if not result.is_finite():
        raise LegacyImportError(f"Expected a finite number at {source}; found {value!r}")
    return result


def _optional_decimal(value: Any) -> Decimal | None:
    try:
        return _decimal_value(value, "optional number")
    except LegacyImportError:
        return None


def _active(value: Any) -> bool:
    return _text(value).casefold() in {"yes", "true", "1", "x", "on"}


def _cell_source(sheet: str, row: int, column: int) -> str:
    return f"{sheet}!{get_column_letter(column)}{row}"


def _require_sheets(workbook: Any) -> None:
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise LegacyImportError(f"Legacy workbook is missing sheet(s): {', '.join(missing)}")


def _method(value: Any, source: str) -> FixingMethod:
    method = METHODS.get(_text(value).casefold())
    if method is None:
        raise LegacyImportError(f"Unknown fixing method at {source}: {value!r}")
    return method


def _currency(unit: str) -> str:
    return "USD" if unit.casefold() in {"bbl", "mmbtu"} else "EUR"


def _underlying_config(
    source_name: str,
    canonical_name: str,
    unit: str,
    fixing_method: FixingMethod,
    active: bool,
) -> UnderlyingConfig:
    delivery_basis = source_name in {"Brent Dated", "HH", "PVB Heren DA (Delivery)"}
    return UnderlyingConfig(
        source_underlying=source_name,
        canonical_underlying=canonical_name,
        fixing_method=fixing_method,
        unit=unit,
        currency=_currency(unit),
        curve_underlying=canonical_name,
        fixing_price_underlying=source_name,
        fixing_price_basis=(
            PriceDateBasis.DELIVERY_DAY if delivery_basis else PriceDateBasis.FIXING_DATE
        ),
        active=active,
        current_month_uses_next_curve=source_name in {"Brent Dated", "HH"},
    )


def _extract_setup(
    workbook: Any, report: LegacyImportReport
) -> tuple[tuple[BookConfig, ...], tuple[UnderlyingConfig, ...]]:
    sheet = workbook["SETUP"]
    books: list[BookConfig] = []
    underlyings: list[UnderlyingConfig] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=3, min_col=1, max_col=16, values_only=True), start=3
    ):
        name = _text(values[1])
        sequence = _optional_decimal(values[0])
        if not name or sequence is None:
            pass
        else:
            books.append(BookConfig(book=name, active=_active(values[2])))
        for first_column, canonical_override in ((7, None), (13, "Index PVB")):
            source_name = _text(values[first_column - 1])
            if not source_name:
                continue
            unit = _text(values[first_column])
            method = _method(
                values[first_column + 1],
                _cell_source("SETUP", row_number, first_column + 2),
            )
            underlyings.append(
                _underlying_config(
                    source_name,
                    canonical_override or source_name,
                    unit,
                    method,
                    _active(values[first_column + 2]),
                )
            )
    if not books or not underlyings:
        raise LegacyImportError("SETUP does not contain the expected BOOK and Underlying lists")
    report.extracted_row_counts["books"] = len(books)
    report.extracted_row_counts["underlyings"] = len(underlyings)
    return tuple(books), tuple(underlyings)


def _extract_config(
    workbook: Any,
    report: LegacyImportReport,
    historical_start: date | None,
    historical_end: date | None,
) -> BuildConfig:
    initial = _date_value(workbook["INITIAL POSITION"]["C2"].value, "INITIAL POSITION!C2")
    start = historical_start or _optional_date(workbook["PROCESS"]["C15"].value) or initial
    end = (
        historical_end
        or _optional_date(workbook["PROCESS"]["D15"].value)
        or _optional_date(workbook["PROCESS"]["C4"].value)
    )
    if end is None:
        raise LegacyImportError(
            "Historical End Date is missing; supply --historical-end or populate PROCESS!D15"
        )
    raw_simulation = _text(workbook["SIMULATION TRADES"]["Q1"].value).upper() or "OFF"
    try:
        simulation = SimulationStatus(raw_simulation)
    except ValueError as exc:
        raise LegacyImportError(
            f"SIMULATION TRADES!Q1 must be ON or OFF; found {raw_simulation!r}"
        ) from exc
    config = BuildConfig(
        initial_market_date=initial,
        historical_start_date=start,
        historical_end_date=end,
        simulation_status=simulation,
    )
    report.initial_market_date = initial.isoformat()
    report.historical_start_date = start.isoformat()
    report.historical_end_date = end.isoformat()
    report.simulation_status = simulation.value
    return config


def _extract_calendar(workbook: Any, report: LegacyImportReport) -> tuple[MarketCalendarDay, ...]:
    sheet = workbook["CALENDAR"]
    rows: list[MarketCalendarDay] = []
    for values in sheet.iter_rows(min_row=4, min_col=1, max_col=5, values_only=True):
        calendar_date = _optional_date(values[0])
        if calendar_date is None:
            continue
        rows.append(
            MarketCalendarDay(
                date=calendar_date,
                is_market_day=_text(values[4]).upper() == "YES",
            )
        )
    if not rows:
        raise LegacyImportError("CALENDAR does not contain any dated rows in A4:E")
    report.extracted_row_counts["calendar"] = len(rows)
    return tuple(rows)


def _extract_initial_exposure(
    workbook: Any, report: LegacyImportReport
) -> tuple[InitialExposure, ...]:
    sheet = workbook["INITIAL POSITION DATA"]
    positions: list[InitialExposure] = []
    zero_rows = 0
    rejected: list[str] = []
    for row, values in enumerate(
        sheet.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True), start=2
    ):
        if all(value in (None, "") for value in values):
            continue
        volume = _optional_decimal(values[4])
        if volume is None:
            rejected.append(str(row))
            continue
        if volume == 0:
            zero_rows += 1
            continue
        try:
            positions.append(
                InitialExposure(
                    initial_market_date=_date_value(values[0], _cell_source(sheet.title, row, 1)),
                    book=_text(values[1]),
                    underlying=_text(values[2]),
                    delivery_month=_date_value(values[3], _cell_source(sheet.title, row, 4)),
                    exposure_volume=volume,
                    source_row_id=f"LEGACY-INITIAL-EXPOSURE-{row}",
                )
            )
        except (LegacyImportError, ValidationError, ValueError):
            rejected.append(str(row))
    if rejected:
        report.add(
            Severity.ERROR,
            "INITIAL_EXPOSURE_ROWS_REJECTED",
            "Non-zero opening rows could not be converted and were excluded.",
            source=SOURCE_RANGES["initial_exposure"],
            count=len(rejected),
            source_rows=tuple(rejected[:25]),
        )
    report.extracted_row_counts["initial_exposure"] = len(positions)
    report.skipped_row_counts["zero_initial_exposure"] = zero_rows
    return tuple(positions)


def _extract_initial_pnl(
    workbook: Any,
    books: tuple[BookConfig, ...],
    report: LegacyImportReport,
) -> tuple[InitialPnl, ...]:
    sheet = workbook["INITIAL POSITION"]
    initial_date = _date_value(sheet["C2"].value, "INITIAL POSITION!C2")
    active_books = {book.book.casefold(): book.book for book in books if book.active}
    rows: list[InitialPnl] = []
    for row, values in enumerate(
        sheet.iter_rows(min_row=5, min_col=1, max_col=3, values_only=True), start=5
    ):
        source_book = _text(values[0])
        canonical_book = active_books.get(source_book.casefold())
        if canonical_book is None:
            continue
        amount = _optional_decimal(values[1])
        if amount is None:
            report.add(
                Severity.ERROR,
                "INITIAL_PNL_ROW_REJECTED",
                "An active BOOK has a blank or non-numeric opening P&L.",
                source=_cell_source(sheet.title, row, 2),
                source_rows=(str(row),),
            )
            continue
        rows.append(
            InitialPnl(
                initial_market_date=initial_date,
                book=canonical_book,
                amount=amount,
                source_row_id=f"LEGACY-INITIAL-PNL-{row}",
                comment=_text(values[2]) or None,
            )
        )
    present = {row.book.casefold() for row in rows}
    missing = sorted(name for key, name in active_books.items() if key not in present)
    if missing:
        report.add(
            Severity.ERROR,
            "INITIAL_PNL_BOOKS_MISSING",
            f"Opening P&L is missing active BOOK(s): {', '.join(missing)}.",
            source=SOURCE_RANGES["initial_pnl"],
            count=len(missing),
        )
    report.extracted_row_counts["initial_pnl"] = len(rows)
    return tuple(rows)


def _extract_trade_sheet(
    workbook: Any,
    sheet_name: str,
    source: TradeSource,
    cutoff: date,
    report: LegacyImportReport,
) -> tuple[Trade, ...]:
    sheet = workbook[sheet_name]
    trades: list[Trade] = []
    rejected: list[str] = []
    zero_qty: list[str] = []
    zero_price: list[str] = []
    after_cutoff: list[str] = []
    comments = 0
    for row, values in enumerate(
        sheet.iter_rows(min_row=3, min_col=1, max_col=15, values_only=True), start=3
    ):
        if all(value in (None, "") for value in values[:9]):
            continue
        try:
            qty = _decimal_value(values[7], _cell_source(sheet_name, row, 8))
            price = _decimal_value(values[8], _cell_source(sheet_name, row, 9))
            trade = Trade(
                source_row_id=f"LEGACY-{source.value}-TRADE-{row}",
                trade_date=_date_value(values[0], _cell_source(sheet_name, row, 1)),
                book=_text(values[1]),
                underlying=_text(values[2]),
                side=Side(_text(values[4]).upper()),
                start_date=_date_value(values[5], _cell_source(sheet_name, row, 6)),
                end_date=_date_value(values[6], _cell_source(sheet_name, row, 7)),
                daily_qty=qty,
                execution_price=price,
                trade_source=source,
                scenario=_text(values[14]) or None if source is TradeSource.SIMULATION else None,
            )
        except (LegacyImportError, ValidationError, ValueError):
            rejected.append(str(row))
            continue
        if trade.trade_date > cutoff:
            after_cutoff.append(str(row))
            continue
        trades.append(trade)
        comments += bool(_text(values[3]))
        if qty == 0:
            zero_qty.append(str(row))
        if price == 0:
            zero_price.append(str(row))
    if rejected:
        report.add(
            Severity.ERROR,
            "TRADE_ROWS_REJECTED",
            f"{sheet_name} rows could not be converted and were excluded.",
            source=f"{sheet_name}!A3:O{sheet.max_row}",
            count=len(rejected),
            source_rows=tuple(rejected[:25]),
        )
    if after_cutoff:
        report.skipped_row_counts[f"{source.value.lower()}_trades_after_cutoff"] = len(after_cutoff)
        report.add(
            Severity.INFO,
            "TRADES_AFTER_CUTOFF_EXCLUDED",
            f"{sheet_name} rows after the {cutoff.isoformat()} as-of date were excluded.",
            source=f"{sheet_name}!A3:A{sheet.max_row}",
            count=len(after_cutoff),
            source_rows=tuple(after_cutoff[:25]),
        )
    if zero_qty:
        report.add(
            Severity.WARNING,
            "ZERO_DAILY_QTY",
            "Explicit zero Daily Qty rows were retained but create no economic event.",
            source=f"{sheet_name}!H3:H{sheet.max_row}",
            count=len(zero_qty),
            source_rows=tuple(zero_qty[:25]),
        )
    if zero_price:
        report.add(
            Severity.WARNING,
            "ZERO_EXECUTION_PRICE",
            "Explicit zero trade prices were retained and require business review.",
            source=f"{sheet_name}!I3:I{sheet.max_row}",
            count=len(zero_price),
            source_rows=tuple(zero_price[:25]),
        )
    if comments:
        report.add(
            Severity.INFO,
            "LEGACY_TRADE_COMMENTS_NOT_IN_ENGINE_SCHEMA",
            "Legacy trade comments remain traceable by source row but are not an economic input.",
            source=f"{sheet_name}!D3:D{sheet.max_row}",
            count=comments,
        )
    return tuple(trades)


CurveGrid = dict[tuple[date, date], tuple[Decimal, str]]
SingleDateCurve = dict[date, tuple[Decimal, str]]


def _curve_grid(
    workbook: Any,
    sheet_name: str,
    *,
    date_column: int,
    header_row: int,
    first_price_column: int,
    data_start_row: int,
    first_market_date: date,
    last_market_date: date,
    zero_is_missing: bool,
    report: LegacyImportReport,
) -> CurveGrid:
    if sheet_name not in workbook.sheetnames:
        report.add(
            Severity.WARNING,
            "OPTIONAL_CURVE_SHEET_MISSING",
            f"Curve source sheet {sheet_name!r} is absent.",
            source=sheet_name,
        )
        return {}
    sheet = workbook[sheet_name]
    header = next(
        sheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        )
    )
    headers = {
        column: delivery_month
        for column, value in enumerate(header, start=1)
        if column >= first_price_column and (delivery_month := _optional_date(value)) is not None
    }
    result: CurveGrid = {}
    zeros = 0
    non_numeric = 0
    duplicates = 0
    for row, values in enumerate(
        sheet.iter_rows(
            min_row=data_start_row,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        ),
        start=data_start_row,
    ):
        market_date = _optional_date(values[date_column - 1])
        if market_date is None or not first_market_date <= market_date <= last_market_date:
            continue
        for column, delivery_month in headers.items():
            raw_value = values[column - 1]
            value = _optional_decimal(raw_value)
            if value is None:
                non_numeric += raw_value not in (None, "")
                continue
            if zero_is_missing and value == 0:
                zeros += 1
                continue
            key = (market_date, delivery_month)
            duplicates += key in result
            result[key] = (value, _cell_source(sheet_name, row, column))
    if zeros:
        report.skipped_row_counts[f"zero_curve_cells_{sheet_name}"] = zeros
    if non_numeric:
        report.add(
            Severity.WARNING,
            "NON_NUMERIC_CURVE_CELLS_SKIPPED",
            f"Non-numeric cached cells in {sheet_name} were not imported as prices.",
            source=sheet_name,
            count=non_numeric,
        )
    if duplicates:
        report.add(
            Severity.WARNING,
            "DUPLICATE_CURVE_KEYS",
            f"Duplicate curve keys in {sheet_name}; the last source row was retained.",
            source=sheet_name,
            count=duplicates,
        )
    return result


def _single_date_curve(
    workbook: Any,
    sheet_name: str,
    *,
    date_column: int,
    price_column: int,
    data_start_row: int,
    data_end_row: int,
    first_market_date: date,
    last_market_date: date,
    report: LegacyImportReport,
) -> SingleDateCurve:
    """Read a dated prompt/spot series used instead of a monthly forward contract."""

    if sheet_name not in workbook.sheetnames:
        report.add(
            Severity.WARNING,
            "OPTIONAL_CURVE_SHEET_MISSING",
            f"Curve source sheet {sheet_name!r} is absent.",
            source=sheet_name,
        )
        return {}
    sheet = workbook[sheet_name]
    result: SingleDateCurve = {}
    zeros = 0
    non_numeric = 0
    duplicates = 0
    for row, values in enumerate(
        sheet.iter_rows(
            min_row=data_start_row,
            max_row=data_end_row,
            min_col=date_column,
            max_col=price_column,
            values_only=True,
        ),
        start=data_start_row,
    ):
        market_date = _optional_date(values[0])
        if market_date is None or not first_market_date <= market_date <= last_market_date:
            continue
        raw_value = values[price_column - date_column]
        value = _optional_decimal(raw_value)
        if value is None:
            non_numeric += raw_value not in (None, "")
            continue
        if value == 0:
            zeros += 1
            continue
        duplicates += market_date in result
        result[market_date] = (value, _cell_source(sheet_name, row, price_column))
    if zeros:
        report.skipped_row_counts[f"zero_prompt_curve_cells_{sheet_name}"] = zeros
    if non_numeric:
        report.add(
            Severity.WARNING,
            "NON_NUMERIC_PROMPT_CURVE_CELLS_SKIPPED",
            f"Non-numeric cached prompt cells in {sheet_name} were not imported as prices.",
            source=sheet_name,
            count=non_numeric,
        )
    if duplicates:
        report.add(
            Severity.WARNING,
            "DUPLICATE_PROMPT_CURVE_KEYS",
            f"Duplicate prompt curve dates in {sheet_name}; the last source row was retained.",
            source=sheet_name,
            count=duplicates,
        )
    return result


def _curve_row(
    market_date: date,
    delivery_month: date,
    underlying: str,
    value: Decimal,
    source: str,
    currency: str,
    unit: str,
) -> CurvePrice:
    return CurvePrice(
        market_date=market_date,
        underlying=underlying,
        delivery_month=delivery_month,
        curve_price=value,
        currency=currency,
        unit=unit,
        source_id=f"LEGACY:{source}",
    )


def _extract_curves(
    workbook: Any, config: BuildConfig, report: LegacyImportReport
) -> tuple[CurvePrice, ...]:
    ttf = _curve_grid(
        workbook,
        "TTF",
        date_column=4,
        header_row=5,
        first_price_column=5,
        data_start_row=9,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        zero_is_missing=True,
        report=report,
    )
    ttf_prompt = _single_date_curve(
        workbook,
        "TTF",
        date_column=45,
        price_column=53,
        data_start_row=7,
        data_end_row=601,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        report=report,
    )
    brent = _curve_grid(
        workbook,
        "Brent Dated",
        date_column=4,
        header_row=5,
        first_price_column=5,
        data_start_row=9,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        zero_is_missing=True,
        report=report,
    )
    hh = _curve_grid(
        workbook,
        "HH",
        date_column=4,
        header_row=5,
        first_price_column=5,
        data_start_row=9,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        zero_is_missing=True,
        report=report,
    )
    pvb_spread = _curve_grid(
        workbook,
        "PVB-TTF",
        date_column=2,
        header_row=2,
        first_price_column=3,
        data_start_row=3,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        zero_is_missing=False,
        report=report,
    )
    peg_spread = _curve_grid(
        workbook,
        "PEG-TTF",
        date_column=2,
        header_row=2,
        first_price_column=3,
        data_start_row=3,
        first_market_date=config.initial_market_date,
        last_market_date=config.historical_end_date,
        zero_is_missing=False,
        report=report,
    )

    rows: dict[tuple[date, str, date], CurvePrice] = {}

    def store(row: CurvePrice) -> None:
        rows[(row.market_date, row.underlying.casefold(), row.delivery_month)] = row

    for (market_date, delivery_month), (value, source) in sorted(ttf.items()):
        if (market_date.year, market_date.month) == (
            delivery_month.year,
            delivery_month.month,
        ):
            continue
        for underlying in ("TTF DA", "TTF MA"):
            store(
                _curve_row(
                    market_date,
                    delivery_month,
                    underlying,
                    value,
                    source,
                    "EUR",
                    "MWh",
                )
            )
        if (spread := pvb_spread.get((market_date, delivery_month))) is not None:
            spread_value, spread_source = spread
            for underlying in ("Index PVB", "Phys PVB", "TVB", "AVB"):
                store(
                    _curve_row(
                        market_date,
                        delivery_month,
                        underlying,
                        value + spread_value,
                        f"{source}+{spread_source}",
                        "EUR",
                        "MWh",
                    )
                )
        if (spread := peg_spread.get((market_date, delivery_month))) is not None:
            spread_value, spread_source = spread
            store(
                _curve_row(
                    market_date,
                    delivery_month,
                    "PEG",
                    value + spread_value,
                    f"{source}+{spread_source}",
                    "EUR",
                    "MWh",
                )
            )

    for market_date, (value, source) in sorted(ttf_prompt.items()):
        # On a month-end Market Date, the dated prompt price applies to the next delivery month.
        delivery_month = (market_date + timedelta(days=1)).replace(day=1)
        for underlying in ("TTF DA", "TTF MA"):
            store(
                _curve_row(
                    market_date,
                    delivery_month,
                    underlying,
                    value,
                    source,
                    "EUR",
                    "MWh",
                )
            )
        if (spread := pvb_spread.get((market_date, delivery_month))) is not None:
            spread_value, spread_source = spread
            for underlying in ("Index PVB", "Phys PVB", "TVB", "AVB"):
                store(
                    _curve_row(
                        market_date,
                        delivery_month,
                        underlying,
                        value + spread_value,
                        f"{source}+{spread_source}",
                        "EUR",
                        "MWh",
                    )
                )
        if (spread := peg_spread.get((market_date, delivery_month))) is not None:
            spread_value, spread_source = spread
            store(
                _curve_row(
                    market_date,
                    delivery_month,
                    "PEG",
                    value + spread_value,
                    f"{source}+{spread_source}",
                    "EUR",
                    "MWh",
                )
            )

    for (market_date, delivery_month), (value, source) in sorted(brent.items()):
        store(
            _curve_row(
                market_date,
                delivery_month,
                "Brent Dated",
                value,
                source,
                "USD",
                "bbl",
            )
        )
    for (market_date, delivery_month), (value, source) in sorted(hh.items()):
        store(_curve_row(market_date, delivery_month, "HH", value, source, "USD", "MMBtu"))
    result = tuple(
        sorted(
            rows.values(),
            key=lambda item: (item.market_date, item.underlying.casefold(), item.delivery_month),
        )
    )
    report.extracted_row_counts["curve_prices"] = len(result)
    return result


def _extract_fixing_prices(
    workbook: Any,
    underlyings: tuple[UnderlyingConfig, ...],
    report: LegacyImportReport,
) -> tuple[FixingPrice, ...]:
    sheet = workbook["FIXING PRICES"]
    profiles = {row.source_underlying.casefold(): row for row in underlyings}
    header = next(
        sheet.iter_rows(
            min_row=4,
            max_row=4,
            min_col=1,
            max_col=sheet.max_column,
            values_only=True,
        )
    )
    result: dict[tuple[date, str], FixingPrice] = {}
    zeros = 0
    unknown_columns: set[str] = set()
    for row, values in enumerate(
        sheet.iter_rows(min_row=6, min_col=1, max_col=sheet.max_column, values_only=True),
        start=6,
    ):
        lookup_date = _optional_date(values[0])
        if lookup_date is None:
            continue
        for column, raw_header in enumerate(header[1:], start=2):
            source_name = _text(raw_header)
            if not source_name:
                continue
            profile = profiles.get(source_name.casefold())
            if profile is None:
                unknown_columns.add(source_name)
                continue
            value = _optional_decimal(values[column - 1])
            if value is None:
                continue
            if value == 0:
                zeros += 1
                continue
            result[(lookup_date, source_name.casefold())] = FixingPrice(
                price_lookup_date=lookup_date,
                underlying=profile.fixing_price_underlying or source_name,
                fixing_price=value,
                currency=profile.currency,
                unit=profile.unit,
                source_id=f"LEGACY:{_cell_source(sheet.title, row, column)}",
            )
    if unknown_columns:
        report.add(
            Severity.WARNING,
            "UNMAPPED_FIXING_PRICE_COLUMNS",
            f"Fixing columns have no active SETUP mapping: {', '.join(sorted(unknown_columns))}.",
            source="FIXING PRICES!B4:S4",
            count=len(unknown_columns),
        )
    report.extracted_row_counts["fixing_prices"] = len(result)
    if zeros:
        report.skipped_row_counts["zero_fixing_price_placeholders"] = zeros
    return tuple(
        sorted(result.values(), key=lambda item: (item.price_lookup_date, item.underlying))
    )


def _add_flow_component(
    totals: dict[tuple[date, str], list[Decimal]],
    market_date: date,
    book: str,
    index: int,
    value: Decimal,
) -> None:
    if value != 0:
        totals[(market_date, book)][index] += value


def _extract_operating_flows(
    workbook: Any, config: BuildConfig, report: LegacyImportReport
) -> tuple[OperatingFlow, ...]:
    totals: dict[tuple[date, str], list[Decimal]] = defaultdict(
        lambda: [Decimal("0"), Decimal("0"), Decimal("0")]
    )
    costs = workbook["COSTS"]
    costs_headers = next(
        costs.iter_rows(min_row=3, max_row=3, min_col=1, max_col=16, values_only=True)
    )
    for values in costs.iter_rows(min_row=5, min_col=1, max_col=16, values_only=True):
        market_date = _optional_date(values[0])
        if (
            market_date is None
            or not config.initial_market_date < market_date <= config.historical_end_date
        ):
            continue
        for column in range(2, 15):
            book = _text(costs_headers[column - 1])
            value = _optional_decimal(values[column - 1])
            if book and value is not None:
                _add_flow_component(totals, market_date, book, 0, value)
        for column in range(15, 17):
            book = _text(costs_headers[column - 1])
            value = _optional_decimal(values[column - 1])
            if book and value is not None:
                _add_flow_component(totals, market_date, book, 1, value)

    foto = workbook["Foto FO"]
    foto_headers = next(
        foto.iter_rows(min_row=4, max_row=4, min_col=1, max_col=22, values_only=True)
    )
    for values in foto.iter_rows(min_row=6, min_col=1, max_col=22, values_only=True):
        market_date = _optional_date(values[1])
        if (
            market_date is None
            or not config.initial_market_date < market_date <= config.historical_end_date
        ):
            continue
        for column in range(18, 20):
            book = _text(foto_headers[column - 1])
            value = _optional_decimal(values[column - 1])
            if book and value is not None:
                _add_flow_component(totals, market_date, book, 1, value)
        for column in range(20, 23):
            book = _text(foto_headers[column - 1])
            value = _optional_decimal(values[column - 1])
            if book and value is not None:
                _add_flow_component(totals, market_date, book, 2, value)

    rows = tuple(
        OperatingFlow(
            market_date=market_date,
            book=book,
            logistics_source_amount=values[0],
            fees_and_optimizations=values[1],
            replication=values[2],
            source_row_id=f"LEGACY-OPERATING-{market_date.isoformat()}-{book}",
        )
        for (market_date, book), values in sorted(totals.items())
        if any(value != 0 for value in values)
    )
    report.extracted_row_counts["operating_flows"] = len(rows)
    return rows


def _add_methodology_issues(bundle: InputBundle, report: LegacyImportReport) -> None:
    registry = {row.source_underlying.casefold(): row for row in bundle.underlyings if row.active}
    relevant = {
        row.underlying.casefold() for row in bundle.initial_exposure if row.exposure_volume != 0
    }
    relevant.update(row.underlying.casefold() for row in bundle.trades if row.daily_qty != 0)
    usd = sorted(
        row.source_underlying
        for key, row in registry.items()
        if key in relevant and row.currency == "USD"
    )
    if usd:
        report.add(
            Severity.ERROR,
            "FX_CONVERSION_NOT_DEFINED",
            (
                "The source contains material USD-priced exposure, but v0.3 does not yet define "
                "the FX conversion into combined EUR P&L. Imported USD prices were not relabelled "
                "as EUR. Do not accept combined P&L until the FX rule is implemented."
            ),
            source="SETUP / Brent Dated / HH / EURF",
            count=len(usd),
            source_rows=tuple(usd),
        )

    available_fixings = {row.underlying.casefold() for row in bundle.fixing_prices}
    missing_series: list[str] = []
    for name in sorted(relevant):
        profile = registry.get(name)
        if profile is None:
            continue
        fixing_name = (profile.fixing_price_underlying or profile.source_underlying).casefold()
        if fixing_name not in available_fixings:
            missing_series.append(profile.fixing_price_underlying or profile.source_underlying)
    if missing_series:
        report.add(
            Severity.ERROR,
            "FIXING_PRICE_SERIES_EMPTY",
            (
                "Material positions use fixing-price series that contain only blanks/zero "
                f"placeholders: {', '.join(sorted(set(missing_series)))}."
            ),
            source=SOURCE_RANGES["fixing_prices"],
            count=len(set(missing_series)),
        )

    report.add(
        Severity.INFO,
        "LEGACY_CALCULATED_OUTPUTS_EXCLUDED",
        (
            "Legacy FIXINGS DATA, EXPOSURE DATA, PNL DATA and related calculated sheets were "
            "not imported as authoritative inputs."
        ),
    )
    report.add(
        Severity.INFO,
        "ENGINE_PREFLIGHT_STILL_REQUIRED",
        (
            "Import success proves transport and schema conversion only. Run excel-build or "
            "build to obtain exact required-price, late-trade and economic validation."
        ),
    )


def read_legacy_workbook(
    workbook_path: str | Path,
    *,
    historical_start: date | None = None,
    historical_end: date | None = None,
) -> tuple[InputBundle, LegacyImportReport]:
    """Extract source inputs without running Excel or changing the legacy workbook."""

    source = Path(workbook_path).expanduser().resolve()
    if not source.exists():
        raise LegacyImportError(f"Legacy workbook does not exist: {source}")
    if source.suffix.casefold() not in {".xlsm", ".xlsx"}:
        raise LegacyImportError("Legacy importer accepts only .xlsm or .xlsx workbooks")
    before = _file_hash(source)
    report = LegacyImportReport(
        source_path=str(source),
        source_sha256=before,
        source_modified_at=datetime.fromtimestamp(source.stat().st_mtime, tz=UTC).isoformat(),
        imported_at=datetime.now(tz=UTC).isoformat(),
    )
    try:
        workbook = load_workbook(
            source,
            data_only=True,
            read_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise LegacyImportError(f"Cannot open legacy workbook headlessly: {exc}") from exc
    try:
        _require_sheets(workbook)
        config = _extract_config(workbook, report, historical_start, historical_end)
        books, underlyings = _extract_setup(workbook, report)
        actual = _extract_trade_sheet(
            workbook,
            "TRADES",
            TradeSource.ACTUAL,
            config.historical_end_date,
            report,
        )
        simulation = _extract_trade_sheet(
            workbook,
            "SIMULATION TRADES",
            TradeSource.SIMULATION,
            config.historical_end_date,
            report,
        )
        bundle = InputBundle(
            config=config,
            books=books,
            underlyings=underlyings,
            calendar=_extract_calendar(workbook, report),
            initial_exposure=_extract_initial_exposure(workbook, report),
            initial_pnl=_extract_initial_pnl(workbook, books, report),
            trades=actual + simulation,
            curve_prices=_extract_curves(workbook, config, report),
            fixing_prices=_extract_fixing_prices(workbook, underlyings, report),
            operating_flows=_extract_operating_flows(workbook, config, report),
            input_hashes={"legacy_workbook": before},
        )
    finally:
        workbook.close()
    after = _file_hash(source)
    if after != before:
        raise LegacyImportError(
            "Legacy workbook changed while it was being read; no output published"
        )
    report.extracted_row_counts["trades"] = len(bundle.trades)
    _add_methodology_issues(bundle, report)
    return bundle, report


def refresh_legacy_curve_table(
    workbook_path: str | Path,
    *,
    historical_end: date | None = None,
) -> tuple[int, date | None]:
    """Rebuild tblCurvePrices in an interface snapshot from its cached provider sheets."""

    source = Path(workbook_path).expanduser().resolve()
    report = LegacyImportReport(
        source_path=str(source),
        source_sha256=_file_hash(source),
        source_modified_at=datetime.fromtimestamp(source.stat().st_mtime, tz=UTC).isoformat(),
        imported_at=datetime.now(tz=UTC).isoformat(),
    )
    cached = load_workbook(source, data_only=True, read_only=True, keep_links=False)
    try:
        manual_dates = cached["MANUAL CHANGES"]
        date_values = {
            _text(row[0]): _optional_date(row[1])
            for row in manual_dates.iter_rows(min_row=6, min_col=1, max_col=2, values_only=True)
            if row[0] not in (None, "")
        }
        initial = date_values.get("Initial Market Date")
        start = date_values.get("Historical Start Date")
        end = historical_end or date_values.get("Historical End Date")
        if initial is None or start is None or end is None:
            raise LegacyImportError("MANUAL CHANGES does not contain the three required dates")
        config = BuildConfig(
            initial_market_date=initial,
            historical_start_date=start,
            historical_end_date=end,
            simulation_status=SimulationStatus.OFF,
        )
        curves = _extract_curves(cached, config, report)
        existing_rows: dict[tuple[date, str, date], tuple[Any, ...]] = {}
        for values in cached["CURVE PRICES"].iter_rows(
            min_row=5, min_col=1, max_col=8, values_only=True
        ):
            market_date = _optional_date(values[0])
            delivery_month = _optional_date(values[2])
            if market_date is None or delivery_month is None or not _text(values[1]):
                continue
            existing_rows[(market_date, _text(values[1]).casefold(), delivery_month)] = values
    finally:
        cached.close()

    merged_rows = existing_rows
    for curve in curves:
        merged_rows[(curve.market_date, curve.underlying.casefold(), curve.delivery_month)] = (
            curve.market_date,
            curve.underlying,
            curve.delivery_month,
            curve.curve_price,
            curve.currency,
            curve.unit,
            curve.source_id,
            curve.source_as_of,
        )
    output_rows = tuple(
        merged_rows[key]
        for key in sorted(merged_rows, key=lambda item: (item[0], item[1], item[2]))
    )

    workbook = load_workbook(source, data_only=False, keep_links=False)
    try:
        sheet = workbook["CURVE PRICES"]
        table = sheet.tables["tblCurvePrices"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for row in sheet.iter_rows(
            min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.value = None
        for row_number, values in enumerate(output_rows, start=min_row + 1):
            for column_number, value in enumerate(values, start=min_col):
                sheet.cell(row_number, column_number, value)
        last_row = min_row + max(1, len(output_rows))
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"
        workbook.save(source)
    finally:
        workbook.close()
    latest = max((curve.market_date for curve in curves), default=None)
    return len(output_rows), latest


def _write_audit_files(report: LegacyImportReport, root: Path) -> tuple[Path, Path]:
    audit_json = root / "legacy_import_audit.json"
    audit_json.write_text(
        json.dumps(report.as_dict(), indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    issues_csv = root / "legacy_import_issues.csv"
    with issues_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("severity", "code", "message", "source", "count", "source_rows"),
            lineterminator="\n",
        )
        writer.writeheader()
        for issue in report.issues:
            row = issue.as_dict()
            row["source_rows"] = ";".join(row["source_rows"])
            writer.writerow(row)
    return audit_json, issues_csv


def _add_import_audit_sheet(path: Path, report: LegacyImportReport) -> None:
    workbook = load_workbook(path, data_only=False, keep_links=False)
    sheet = workbook.create_sheet("LEGACY IMPORT", 1)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "C55A11"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "LEGACY IMPORT AUDIT"
    sheet["A1"].fill = PatternFill("solid", fgColor="843C0C")
    sheet["A1"].font = Font(name="Aptos Display", size=18, color="FFFFFF", bold=True)
    sheet["A2"] = "Import status"
    sheet["B2"] = report.status
    sheet["A3"] = "Source workbook"
    sheet["B3"] = report.source_path
    sheet["A4"] = "Source SHA-256"
    sheet["B4"] = report.source_sha256
    sheet["A5"] = "Imported at (UTC)"
    sheet["B5"] = report.imported_at
    sheet["A6"] = "Errors requiring review"
    sheet["B6"] = report.error_count
    sheet["A7"] = "Warnings"
    sheet["B7"] = report.warning_count
    for row in range(2, 8):
        sheet.cell(row, 1).font = Font(name="Aptos", bold=True)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    headers = ("Severity", "Code", "Message", "Source", "Count", "Source rows")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(10, column, header)
        cell.fill = PatternFill("solid", fgColor="C55A11")
        cell.font = Font(name="Aptos", color="FFFFFF", bold=True)
    issues = report.issues or [ImportIssue(Severity.INFO, "NONE", "No import issues.")]
    for row, issue in enumerate(issues, start=11):
        values = (
            issue.severity.value,
            issue.code,
            issue.message,
            issue.source,
            issue.count,
            "; ".join(issue.source_rows),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")
    last_row = 10 + len(issues)
    table = Table(displayName="tblLegacyImportIssues", ref=f"A10:F{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A11"
    for column, width in enumerate((12, 34, 80, 40, 10, 40), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    workbook["START HERE"]["B4"] = "IMPORTED — NOT BUILT"
    workbook["START HERE"]["B4"].fill = PatternFill("solid", fgColor="FFF2CC")
    workbook.save(path)
    workbook.close()


def import_legacy_workbook(
    workbook_path: str | Path,
    output_directory: str | Path,
    *,
    historical_start: date | None = None,
    historical_end: date | None = None,
) -> LegacyImportResult:
    """Create normalized and Excel v0.3 inputs from one legacy workbook snapshot."""

    destination = Path(output_directory).expanduser().resolve()
    if destination.exists():
        raise LegacyImportError(f"Import output already exists: {destination}")
    bundle, report = read_legacy_workbook(
        workbook_path,
        historical_start=historical_start,
        historical_end=historical_end,
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    staging = Path(tempfile.mkdtemp(prefix=f".{destination.name}-", dir=destination.parent))
    try:
        normalized = write_bundle(bundle, staging / "normalized_bundle")
        excel = create_excel_template(staging / "GTM_Imported_Input.xlsx", bundle=bundle)
        _add_import_audit_sheet(excel, report)
        audit_json, issues_csv = _write_audit_files(report, staging)
        os.replace(staging, destination)
        destination.chmod(0o755)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    return LegacyImportResult(
        bundle=bundle,
        report=report,
        output_directory=destination,
        normalized_bundle=destination / normalized.relative_to(staging),
        excel_workbook=destination / excel.relative_to(staging),
        audit_json=destination / audit_json.relative_to(staging),
        issues_csv=destination / issues_csv.relative_to(staging),
    )
