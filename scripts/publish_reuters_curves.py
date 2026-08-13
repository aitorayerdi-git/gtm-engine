"""Publish cached Reuters staging matrices into normalized CURVE PRICES and FX RATES."""
from __future__ import annotations

import argparse
import csv
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def number(value: Any) -> Decimal | None:
    try:
        result = Decimal(str(value))
        return result if result.is_finite() else None
    except (InvalidOperation, TypeError, ValueError):
        return None


def grid(ws: Any, date_col: int, header_row: int, first_col: int, first_row: int, start: date, end: date, zero_missing: bool) -> dict[tuple[date,date], Decimal]:
    header_values = next(ws.iter_rows(min_row=header_row,max_row=header_row,values_only=True))
    headers = {c: as_date(header_values[c-1]) for c in range(first_col,ws.max_column+1)}
    headers = {c:d for c,d in headers.items() if d is not None}
    out: dict[tuple[date,date],Decimal] = {}
    for row in ws.iter_rows(min_row=first_row,values_only=True):
        md=as_date(row[date_col-1])
        if md is None or not start <= md <= end: continue
        for c,dm in headers.items():
            v=number(row[c-1])
            if v is not None and not (zero_missing and v==0): out[(md,dm)]=v
    return out


def carry_forward_missing(
    values: dict[tuple[date, date], Decimal], target_dates: tuple[date, ...]
) -> None:
    delivery_months = {delivery_month for _, delivery_month in values}
    for target in target_dates:
        for delivery_month in delivery_months:
            key = (target, delivery_month)
            if key in values:
                continue
            prior_dates = [
                market_date
                for market_date, month in values
                if month == delivery_month and market_date < target
            ]
            if prior_dates:
                values[key] = values[(max(prior_dates), delivery_month)]


def main() -> int:
    parser=argparse.ArgumentParser(); parser.add_argument('workbook',type=Path); parser.add_argument('--export-dir',type=Path); args=parser.parse_args()
    path=args.workbook.resolve()
    values=load_workbook(path,data_only=True,read_only=True,keep_links=False)
    formulas=None
    try:
        manual=values['MANUAL CHANGES']; start=as_date(manual['P4'].value); end=as_date(manual['P6'].value)
        if start is None or end is None: raise ValueError('MANUAL dates are missing')
        ttf=grid(values['TTF'],4,5,5,9,start,end,True); brent=grid(values['Brent Dated'],4,5,5,9,start,end,True); hh=grid(values['HH'],4,5,5,9,start,end,True)
        pvb=grid(values['PVB-TTF'],2,2,3,3,start,end,False); peg=grid(values['PEG-TTF'],2,2,3,3,start,end,False)
        # Reviewed source exceptions: retain the immediately preceding available snapshot.
        carry_forward_missing(hh,(date(2026,7,30),date(2026,7,31)))
        carry_forward_missing(pvb,(date(2026,7,22),date(2026,8,6)))
        carry_forward_missing(peg,(date(2026,7,22),date(2026,8,6)))
        rows: dict[tuple[date,str,date],tuple[Any,...]]={}
        def put(md:date,u:str,dm:date,v:Decimal,ccy:str,unit:str,src:str)->None: rows[(md,u,dm)]=(md,u,dm,float(v),ccy,unit,src,datetime.now())
        for (md,dm),v in ttf.items():
            if (md.year,md.month)==(dm.year,dm.month): continue
            for u in ('TTF DA','TTF MA'): put(md,u,dm,v,'EUR','MWh','REUTERS:TTF')
            if (md,dm) in pvb:
                for u in ('Index PVB','Phys PVB','TVB','AVB'): put(md,u,dm,v+pvb[(md,dm)],'EUR','MWh','REUTERS:TTF+PVB-TTF')
            if (md,dm) in peg: put(md,'PEG',dm,v+peg[(md,dm)],'EUR','MWh','REUTERS:TTF+PEG-TTF')
        # Month-end prompt TTF override (AS=date, BA=price).
        ws=values['TTF']
        for row in ws.iter_rows(min_row=7,min_col=45,max_col=53,values_only=True):
            md=as_date(row[0]); v=number(row[8])
            if md is None or v is None or v==0 or not start<=md<=end: continue
            dm=(md+timedelta(days=1)).replace(day=1)
            for u in ('TTF DA','TTF MA'): put(md,u,dm,v,'EUR','MWh','REUTERS:TTF-PROMPT')
            if (md,dm) in pvb:
                for u in ('Index PVB','Phys PVB','TVB','AVB'): put(md,u,dm,v+pvb[(md,dm)],'EUR','MWh','REUTERS:TTF-PROMPT+PVB-TTF')
            if (md,dm) in peg: put(md,'PEG',dm,v+peg[(md,dm)],'EUR','MWh','REUTERS:TTF-PROMPT+PEG-TTF')
        for (md,dm),v in brent.items(): put(md,'Brent Dated',dm,v,'USD','bbl','REUTERS:BRENT')
        for (md,dm),v in hh.items(): put(md,'HH',dm,v,'USD','MMBtu','REUTERS:HH')
        existing_rows={}
        for row in values['CURVE PRICES'].iter_rows(min_row=5,max_col=8,values_only=True):
            md=as_date(row[0])
            dm=as_date(row[2])
            if md is not None and dm is not None: existing_rows[(md,str(row[1]),dm)]=(md,row[1],dm,*row[3:])
        existing_rows.update(rows)
        published=[existing_rows[k] for k in sorted(existing_rows)]
        fx_by_key={}
        for row in values['FX RATES'].iter_rows(min_row=5,max_col=5,values_only=True):
            rd=as_date(row[0])
            if rd is not None: fx_by_key[(rd,str(row[1]).upper())]=(rd,*row[1:])
        eurf=values['EURF']
        for row in eurf.iter_rows(min_row=8,min_col=4,max_col=5,values_only=True):
            rd=as_date(row[0]); rate=number(row[1])
            if rd is not None and rate is not None and rate!=0 and start<=rd<=end: fx_by_key[(rd,'USD')]=(rd,'USD',float(rate),'REUTERS:EURF-SPOT',datetime.now())
        fxrows=list(fx_by_key.values())
        fxrows.sort(key=lambda x:(x[0],str(x[1])))
        if args.export_dir:
            args.export_dir.mkdir(parents=True,exist_ok=True)
            for name,data in (('curve_prices.csv',published),('fx_rates.csv',fxrows)):
                with (args.export_dir/name).open('w',newline='',encoding='utf-8-sig') as handle:
                    writer=csv.writer(handle); writer.writerows(data)
            print({'curve_prices_total':len(published),'curve_prices_published':len(rows),'fx_rates_total':len(fxrows),'through':end.isoformat()})
            return 0
        formulas=load_workbook(path,data_only=False,read_only=False,keep_vba=True,keep_links=True)
        cp=formulas['CURVE PRICES']; table=cp.tables['tblCurvePrices']
        cp.delete_rows(5,max(1,cp.max_row-4))
        for r,row in enumerate(published,5):
            for c,v in enumerate(row,1): cp.cell(r,c,v)
        table.ref=f'A4:H{max(5,4+len(published))}'
        fx=formulas['FX RATES']; fxt=fx.tables['tblFxRates']
        fx.delete_rows(5,max(1,fx.max_row-4))
        for r,row in enumerate(fxrows,5):
            for c,v in enumerate(row,1): fx.cell(r,c,v)
        fxt.ref=f'A4:E{max(5,4+len(fxrows))}'
        formulas.save(path)
        print({'curve_prices_published':len(rows),'fx_rates_total':len(fxrows),'through':end.isoformat()})
    finally:
        values.close()
        if formulas is not None: formulas.close()
    return 0


if __name__=='__main__': raise SystemExit(main())
