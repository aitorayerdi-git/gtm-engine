"""Create one active-sheet workbook per sheet for headless visual QA."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    source_workbook = load_workbook(args.source, read_only=True)
    names = source_workbook.sheetnames
    source_workbook.close()
    for index, name in enumerate(names):
        workbook = load_workbook(args.source, data_only=False, keep_links=False)
        for sheet in workbook.worksheets:
            sheet.sheet_state = "hidden"
        workbook[name].sheet_state = "visible"
        workbook.active = index
        slug = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
        destination = args.output / f"{index + 1:02d}_{slug}.xlsx"
        workbook.save(destination)
        workbook.close()
        print(destination)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
