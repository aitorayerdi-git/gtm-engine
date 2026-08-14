"""Regenerate the reviewed fixing-methodology input and summary workbooks."""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from gtm_engine.canonicalize import Registry
from gtm_engine.excel import create_excel_template, load_excel_bundle, load_setup_mapping
from gtm_engine.fx import FxIndex
from gtm_engine.models import (
    BuildConfig,
    BuildStatus,
    CurvePrice,
    DeliveryElection,
    FixingMethod,
    FixingPrice,
    FxRate,
    InitialPnl,
    InputBundle,
    MarketCalendarDay,
    Side,
    Trade,
    TradeSource,
)
from gtm_engine.pipeline import build

ZERO = Decimal("0")
DELIVERY_COLUMNS = {
    "TTFDA Heren": "TTFDA Delivery",
    "PVB Heren": "PVB Heren Delivery",
}


def _days(first: date, last: date) -> tuple[date, ...]:
    return tuple(first + timedelta(days=i) for i in range((last - first).days + 1))


def _delivery_elections(enabled: bool) -> tuple[DeliveryElection, ...]:
    if not enabled:
        return ()
    return (
        DeliveryElection(
            decision_date=date(2026, 6, 30),
            book="CGTO",
            underlying="TTFDA Heren",
            side=Side.BUY,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            delivery_daily_qty=Decimal("10"),
            unit="MWh",
            source_row_id="TEST-DELIVERY-TTFDA-JUL26",
            comment="All July 2026 TTFDA Heren volume elected for delivery.",
        ),
        DeliveryElection(
            decision_date=date(2026, 6, 30),
            book="CGTO",
            underlying="PVB Heren",
            side=Side.BUY,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            delivery_daily_qty=Decimal("10"),
            unit="MWh",
            source_row_id="TEST-DELIVERY-PVB-JUL26",
            comment="All July 2026 PVB Heren volume elected for delivery.",
        ),
    )


def _synthetic_bundle(mapping: Path, *, july_delivery: bool = False) -> InputBundle:
    books, mapped = load_setup_mapping(mapping)
    underlyings = tuple(
        row.model_copy(update={"currency": "USD"})
        if row.source_underlying in {"Brent Dated", "HH"}
        else row
        for row in mapped
    )
    config = BuildConfig(
        initial_market_date=date(2025, 12, 30),
        historical_start_date=date(2025, 12, 31),
        historical_end_date=date(2026, 7, 10),
    )
    calendar = tuple(
        MarketCalendarDay(date=value, is_market_day=value.weekday() < 5)
        for value in _days(date(2025, 12, 1), date(2028, 12, 31))
    )
    trades = tuple(
        Trade(
            source_row_id=f"TEST-TRADE-{index:02d}",
            trade_date=date(2025, 12, 31),
            book="CGTO",
            underlying=profile.source_underlying,
            side=Side.BUY,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            daily_qty=Decimal("10"),
            execution_price=Decimal("100"),
            trade_source=TradeSource.ACTUAL,
        )
        for index, profile in enumerate(underlyings, start=1)
    )
    fixing_profiles = {profile.fixing_price_underlying: profile for profile in underlyings}
    profile_order = {name: index for index, name in enumerate(sorted(fixing_profiles), start=1)}
    fixing_prices: list[FixingPrice] = []
    for calendar_day in calendar:
        for name, profile in sorted(fixing_profiles.items()):
            if name == "TTFDA Heren":
                value = Decimal("50")
            elif name == "PVB Heren":
                value = Decimal("70")
            elif name == "Brent Dated":
                value = Decimal("80")
            elif name == "HH":
                value = Decimal("4")
            elif profile.fixing_method is FixingMethod.MONTH_AHEAD:
                value = Decimal("30") + Decimal(calendar_day.date.timetuple().tm_yday) / Decimal(
                    "1000"
                )
            else:
                value = Decimal("10") + Decimal(profile_order[name]) / Decimal("10")
            fixing_prices.append(
                FixingPrice(
                    price_lookup_date=calendar_day.date,
                    underlying=name,
                    fixing_price=value,
                    currency=profile.currency,
                    unit=profile.unit,
                    source_id=f"SYNTH-FIX-{name}-{calendar_day.date.isoformat()}",
                )
            )
    curve_profiles = {profile.curve_underlying: profile for profile in underlyings}
    delivery_months = tuple(date(2026, month, 1) for month in range(1, 13))
    curve_prices: list[CurvePrice] = []
    for calendar_day in calendar:
        if not calendar_day.is_market_day or not (
            config.initial_market_date <= calendar_day.date <= config.historical_end_date
        ):
            continue
        for name, profile in sorted(curve_profiles.items()):
            value = (
                Decimal("80")
                if name == "Brent Dated"
                else Decimal("4")
                if name == "HH"
                else Decimal("100")
            )
            for delivery_month in delivery_months:
                curve_prices.append(
                    CurvePrice(
                        market_date=calendar_day.date,
                        underlying=name,
                        delivery_month=delivery_month,
                        curve_price=value,
                        currency=profile.currency,
                        unit=profile.unit,
                        source_id=(
                            f"SYNTH-CURVE-{name}-{calendar_day.date.isoformat()}-"
                            f"{delivery_month.isoformat()}"
                        ),
                    )
                )
    fx_rates = tuple(
        FxRate(
            rate_date=row.date,
            currency="USD",
            currency_per_eur=Decimal("1.20"),
            source_id=f"SYNTH-FX-USD-{row.date.isoformat()}",
        )
        for row in calendar
        if row.is_market_day and row.date <= config.historical_end_date
    )
    return InputBundle(
        config=config,
        books=books,
        underlyings=underlyings,
        calendar=calendar,
        initial_exposure=(),
        initial_pnl=tuple(
            InitialPnl(
                initial_market_date=config.initial_market_date,
                book=book.book,
                amount=ZERO,
                source_row_id=f"SYNTH-IPNL-{index:02d}",
            )
            for index, book in enumerate(books, start=1)
        ),
        trades=trades,
        delivery_elections=_delivery_elections(july_delivery),
        curve_prices=tuple(curve_prices),
        fixing_prices=tuple(fixing_prices),
        fx_rates=fx_rates,
        operating_flows=(),
    )


def _updated_bundle(source: Path, mapping: Path, *, july_delivery: bool = False):
    bundle = load_excel_bundle(source)
    _books, mapped = load_setup_mapping(mapping)
    existing = {row.source_underlying: row for row in bundle.underlyings}
    underlyings = tuple(
        row.model_copy(update={"currency": existing[row.source_underlying].currency})
        if row.source_underlying in existing
        else row
        for row in mapped
    )

    trade_names = {
        "TTF DA": "TTFDA Heren",
        "PVB Heren DA": "PVB Heren",
    }
    delivery_trade_names = {
        "TTF DA (Delivery)",
        "TTFDA Delivery",
        "PVB Heren DA (Delivery)",
        "PVB Heren Delivery",
    }
    trades: list[Trade] = []
    for row in bundle.trades:
        if row.underlying in delivery_trade_names:
            continue
        trades.append(
            row.model_copy(update={"underlying": trade_names.get(row.underlying, row.underlying)})
        )

    ordinary_prices: list[FixingPrice] = []
    selected_prices: dict[tuple[date, str], FixingPrice] = {}
    for row in bundle.fixing_prices:
        if row.underlying in {"TTF DA (Delivery)", "TTFDA Heren"}:
            selected_prices[(row.price_lookup_date, "TTFDA Heren")] = row.model_copy(
                update={"underlying": "TTFDA Heren"}
            )
        elif row.underlying in {"PVB Heren DA (Delivery)", "PVB Heren"}:
            selected_prices[(row.price_lookup_date, "PVB Heren")] = row.model_copy(
                update={"underlying": "PVB Heren"}
            )
        elif row.underlying not in {
            "TTF DA",
            "TTFDA Delivery",
            "PVB Heren DA",
            "PVB Heren Delivery",
        }:
            ordinary_prices.append(row)
    prices = [*ordinary_prices, *selected_prices.values()]
    prices.sort(key=lambda row: (row.price_lookup_date, row.underlying, row.source_id))
    delivery_elections = _delivery_elections(july_delivery)
    return bundle.model_copy(
        update={
            "underlyings": underlyings,
            "trades": tuple(trades),
            "delivery_elections": delivery_elections,
            "fixing_prices": tuple(prices),
            "input_hashes": {},
        }
    )


def _write_matrix(
    workbook: Workbook,
    title: str,
    date_heading: str,
    dates: tuple[date, ...],
    columns: tuple[str, ...],
    values: dict[tuple[date, str], Decimal],
) -> None:
    sheet = workbook.create_sheet(title)
    headers = (date_heading, *columns)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(color="FFFFFF", bold=True)
    for axis_date in dates:
        sheet.append((axis_date, *(float(values[(axis_date, name)]) for name in columns)))
    sheet.freeze_panes = "B2"
    table = Table(
        displayName="tbl"
        + "".join(character for character in title.title() if character.isalnum()),
        ref=f"A1:{get_column_letter(len(headers))}{len(dates) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    sheet.column_dimensions["A"].width = 15
    for column in range(2, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    for cell in sheet["A"][1:]:
        cell.number_format = "dd-mmm-yyyy"
    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "#,##0.00;[Red]-#,##0.00;0.00"


def _write_exposure_sheets(workbook: Workbook, result) -> None:
    exposure_by_key = {
        (
            row.market_date,
            row.book,
            row.underlying,
            row.delivery_month,
            row.trade_source,
            row.scenario,
        ): row
        for row in result.exposure
    }
    headers = (
        "Market Date",
        "Previous Market Date",
        "BOOK",
        "Underlying",
        "Delivery Month",
        "Trade Source",
        "Scenario",
        "Exposure Volume",
        "Curve Price",
        "Exposure MtM",
        "Gross Delta Exposure MtM",
        "Trade Entry Adjustment",
        "Delta Exposure MtM",
        "Currency",
        "Explicit Closure",
        "Build ID",
    )
    detail = workbook.create_sheet("Exposure data")
    detail.append(headers)
    for cell in detail[1]:
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(color="FFFFFF", bold=True)
    pnl_rows = [row for row in result.pnl if row.delivery_month is not None]
    for pnl in pnl_rows:
        key = (
            pnl.market_date,
            pnl.book,
            pnl.underlying,
            pnl.delivery_month,
            pnl.trade_source,
            pnl.scenario,
        )
        exposure = exposure_by_key.get(key)
        detail.append(
            (
                pnl.market_date,
                pnl.previous_market_date,
                pnl.book,
                pnl.underlying,
                pnl.delivery_month,
                pnl.trade_source.value if pnl.trade_source else None,
                pnl.scenario,
                float(exposure.exposure_volume) if exposure else 0,
                float(exposure.curve_price) if exposure and exposure.curve_price is not None else 0,
                float(pnl.exposure_mtm),
                float(pnl.gross_delta_exposure_mtm),
                float(pnl.trade_entry_adjustment),
                float(pnl.delta_exposure_mtm),
                exposure.currency if exposure else "EUR",
                exposure.is_explicit_closure if exposure else False,
                pnl.build_id,
            )
        )
    table = Table(displayName="tblExposureData", ref=f"A1:P{detail.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    detail.add_table(table)
    detail.freeze_panes = "A2"
    for index, width in enumerate(
        (14, 20, 18, 24, 16, 16, 16, 18, 16, 18, 26, 24, 22, 12, 18, 38), start=1
    ):
        detail.column_dimensions[get_column_letter(index)].width = width
    for row in detail.iter_rows(min_row=2):
        for index in (1, 2, 5):
            row[index - 1].number_format = "dd-mmm-yyyy"
        for index in range(8, 14):
            row[index - 1].number_format = "#,##0.00;[Red]-#,##0.00;0.00"

    market_dates = sorted({row.market_date for row in pnl_rows})
    books = sorted({row.book for row in pnl_rows})
    underlyings = sorted({row.underlying for row in pnl_rows})
    lists = workbook.create_sheet("Lists")
    lists.append(("Market Dates", "BOOKS", "Underlyings"))
    for index, value in enumerate(market_dates, start=2):
        lists.cell(index, 1, value).number_format = "dd-mmm-yyyy"
    for index, value in enumerate(books, start=2):
        lists.cell(index, 2, value)
    for index, value in enumerate(underlyings, start=2):
        lists.cell(index, 3, value)
    workbook.defined_names.add(
        DefinedName("ExposureMarketDates", attr_text=f"'Lists'!$A$2:$A${len(market_dates) + 1}")
    )
    workbook.defined_names.add(
        DefinedName("ExposureBooks", attr_text=f"'Lists'!$B$2:$B${len(books) + 1}")
    )
    lists.sheet_state = "hidden"

    delivery_months = tuple(
        date(year, month, 1) for year in range(2026, 2029) for month in range(1, 13)
    )

    def add_selector_sheet(title: str, detail_column: int) -> None:
        sheet = workbook.create_sheet(title)
        sheet["A1"] = title
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
        sheet["A2"], sheet["B2"] = "Market Date", max(market_dates)
        sheet["B2"].number_format = "dd-mmm-yyyy"
        sheet["A3"], sheet["B3"] = "BOOK", "CGTO" if "CGTO" in books else books[0]
        for cell in (sheet["A2"], sheet["A3"]):
            cell.font = Font(bold=True)
        for cell in (sheet["B2"], sheet["B3"]):
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.font = Font(color="0000FF")
        date_validation = DataValidation(
            type="list", formula1="=ExposureMarketDates", allow_blank=False
        )
        book_validation = DataValidation(type="list", formula1="=ExposureBooks", allow_blank=False)
        sheet.add_data_validation(date_validation)
        sheet.add_data_validation(book_validation)
        date_validation.add(sheet["B2"])
        book_validation.add(sheet["B3"])
        view_headers = ("Delivery Month", *underlyings)
        for column, value in enumerate(view_headers, start=1):
            sheet.cell(5, column, value)
        for cell in sheet[5]:
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(color="FFFFFF", bold=True)
        last_detail = detail.max_row
        letter = get_column_letter(detail_column)
        for row_number, delivery_month in enumerate(delivery_months, start=6):
            sheet.cell(row_number, 1, delivery_month).number_format = "mmm-yyyy"
            for column_number, _underlying_name in enumerate(underlyings, start=2):
                formula = (
                    f"=SUMIFS('Exposure data'!${letter}$2:${letter}${last_detail},"
                    f"'Exposure data'!$A$2:$A${last_detail},$B$2,"
                    f"'Exposure data'!$C$2:$C${last_detail},$B$3,"
                    f"'Exposure data'!$E$2:$E${last_detail},$A{row_number},"
                    f"'Exposure data'!$D$2:$D${last_detail},{get_column_letter(column_number)}$5)"
                )
                cell = sheet.cell(row_number, column_number, formula)
                cell.number_format = "#,##0.00;[Red]-#,##0.00;0.00"
        sheet.freeze_panes = "B6"
        sheet.column_dimensions["A"].width = 18
        for column in range(2, len(view_headers) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
        last_cell = f"{get_column_letter(len(view_headers))}{len(delivery_months) + 5}"
        sheet.auto_filter.ref = f"A5:{last_cell}"
        sheet.conditional_formatting.add(
            f"B6:{last_cell}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(color="9C0006"),
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    add_selector_sheet("Exposure by Market Date", 8)
    add_selector_sheet("Delta Exposure MtM", 13)


def write_business_output(destination: Path, bundle, result) -> None:
    """Write the reviewed seven-sheet fixing and exposure business output."""
    columns_list: list[str] = []
    for profile in bundle.underlyings:
        if not profile.active:
            continue
        columns_list.append(profile.source_underlying)
        delivery_name = DELIVERY_COLUMNS.get(profile.source_underlying)
        if delivery_name is not None:
            columns_list.append(delivery_name)
    columns = tuple(columns_list)
    calendar_dates = _days(bundle.config.historical_start_date, bundle.config.historical_end_date)
    market_dates = tuple(
        row.date
        for row in bundle.calendar
        if row.is_market_day
        and bundle.config.historical_start_date <= row.date <= bundle.config.historical_end_date
    )
    volume: dict[tuple[date, str], Decimal] = defaultdict(lambda: ZERO)
    pnl_fixing_date: dict[tuple[date, str], Decimal] = defaultdict(lambda: ZERO)
    pnl_market_date: dict[tuple[date, str], Decimal] = defaultdict(lambda: ZERO)
    registry = Registry.from_bundle(bundle)
    fx = FxIndex(bundle.fx_rates)
    for row in result.fixings:
        volume[(row.fixing_date, row.source_underlying)] += row.fixing_volume
        base_name = next(
            (
                base
                for base, delivery in DELIVERY_COLUMNS.items()
                if delivery == row.source_underlying
            ),
            row.source_underlying,
        )
        profile = registry.underlying(base_name)
        include_in_pnl = row.source_underlying in DELIVERY_COLUMNS.values() or (
            profile is not None and profile.include_fixing_in_pnl
        )
        if include_in_pnl:
            economic = -fx.to_eur(row.fixing_amount, row.currency, row.fixing_date)
            pnl_fixing_date[(row.fixing_date, row.source_underlying)] += economic
            pnl_market_date[(row.applied_market_date, row.source_underlying)] += economic

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_matrix(
        workbook,
        "Fixing volume by fixing date",
        "Fixing Date",
        calendar_dates,
        columns,
        volume,
    )
    _write_matrix(
        workbook,
        "Fixing PnL by fixing date",
        "Fixing Date",
        calendar_dates,
        columns,
        pnl_fixing_date,
    )
    _write_matrix(
        workbook,
        "Fixing PnL by market date",
        "Market Date",
        market_dates,
        columns,
        pnl_market_date,
    )
    _write_exposure_sheets(workbook, result)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        type=Path,
        help="Optional prior input workbook to normalize; omit for a self-contained synthetic run.",
    )
    parser.add_argument(
        "--mapping", type=Path, default=Path("docs/GTM_ACTIVE_SETUP_MAPPING_v0.3.csv")
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--july-delivery",
        action="store_true",
        help="Elect all July 2026 TTFDA Heren and PVB Heren BUY volume for delivery.",
    )
    args = parser.parse_args()
    bundle = (
        _updated_bundle(args.source, args.mapping, july_delivery=args.july_delivery)
        if args.source is not None
        else _synthetic_bundle(args.mapping, july_delivery=args.july_delivery)
    )
    result = build(bundle)
    if result.manifest.status is not BuildStatus.VERIFIED:
        errors = [row.model_dump(mode="json") for row in result.validation]
        raise RuntimeError(f"Fixing test build did not verify: {errors}")
    create_excel_template(args.input, bundle=bundle)
    roundtrip_bundle = load_excel_bundle(args.input)
    roundtrip_result = build(roundtrip_bundle)
    if roundtrip_result.manifest.status is not BuildStatus.VERIFIED:
        errors = [row.model_dump(mode="json") for row in roundtrip_result.validation]
        raise RuntimeError(f"Generated Excel did not verify after reload: {errors}")
    write_business_output(args.output, roundtrip_bundle, roundtrip_result)
    delivery_totals = {
        name: sum(
            (
                row.fixing_volume
                for row in roundtrip_result.fixings
                if row.source_underlying == name
            ),
            ZERO,
        )
        for name in DELIVERY_COLUMNS.values()
    }
    print(
        {
            "status": roundtrip_result.manifest.status.value,
            "trades": len(roundtrip_bundle.trades),
            "delivery_elections": len(roundtrip_bundle.delivery_elections),
            "delivery_fixing_volume": delivery_totals,
            "input": str(args.input),
            "output": str(args.output),
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
